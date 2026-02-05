import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import numpy as np
from io import BytesIO
from datetime import datetime

# ═══════════════════════════════════════════════════════════════════════════════
# CONFIG
# ═══════════════════════════════════════════════════════════════════════════════
st.set_page_config(page_title="Solutions Pipeline | Marken", page_icon="📋", layout="wide", initial_sidebar_state="expanded")

# ── Marken Brand Palette ─────────────────────────────────────────────────────
NY    = "#002B49"    # Marken navy
NY2   = "#003A63"
TL    = "#00857C"    # Marken teal
TLL   = "#E8F5F3"
GD    = "#FFB500"    # UPS gold
W     = "#FFFFFF"
G50   = "#FAFBFC"
G100  = "#F4F5F7"
G200  = "#E1E4E8"
G400  = "#A0A8B4"
G600  = "#6B7280"
G800  = "#2D3748"
BA    = "#2E86AB"
RD    = "#DC3545"
GN    = "#28A745"
OR    = "#E06C47"
SEQ   = [NY, TL, GD, BA, "#6C5B7B", OR, GN, "#9B59B6", "#F39C12", "#1ABC9C"]

# ── CSS ──────────────────────────────────────────────────────────────────────
st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;600;700&family=Instrument+Serif&display=swap');
.stApp {{ background:{G50}; font-family:'DM Sans',sans-serif; color:{G800}; }}
h1,h2,h3,h4 {{ font-family:'DM Sans',sans-serif; color:{NY}; }}

/* Sidebar */
section[data-testid="stSidebar"] {{ background:{NY}; }}
section[data-testid="stSidebar"] * {{ color:{W} !important; font-family:'DM Sans',sans-serif; }}
section[data-testid="stSidebar"] hr {{ border-color:rgba(255,255,255,0.12); }}

/* Header */
.rh {{ background:{NY}; padding:1.5rem 2rem 1.3rem; border-radius:6px; margin-bottom:1.1rem;
    display:flex; justify-content:space-between; align-items:flex-end; }}
.rh-b {{ font-size:.6rem; font-weight:600; letter-spacing:.22em; text-transform:uppercase; color:{GD}; margin-bottom:.3rem; }}
.rh-t {{ font-family:'Instrument Serif',Georgia,serif; font-size:1.45rem; color:{W}; line-height:1.2; }}
.rh-s {{ font-size:.76rem; color:rgba(255,255,255,0.5); margin-top:.2rem; }}
.rh-r {{ text-align:right; }}
.rh-d {{ font-size:.7rem; color:rgba(255,255,255,0.45); }}
.rh-c {{ font-size:.56rem; color:{GD}; font-weight:600; letter-spacing:.12em; text-transform:uppercase; margin-top:.12rem; }}

/* KPIs */
.kr {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(150px,1fr)); gap:.7rem; margin-bottom:1.1rem; }}
.kp {{ background:{W}; border:1px solid {G200}; border-top:3px solid var(--ac,{TL}); border-radius:4px; padding:.9rem 1rem .8rem; }}
.kp-l {{ font-size:.6rem; font-weight:600; letter-spacing:.1em; text-transform:uppercase; color:{G400}; margin-bottom:.25rem; }}
.kp-v {{ font-size:1.45rem; font-weight:700; color:{NY}; letter-spacing:-.03em; line-height:1.1; }}
.kp-d {{ font-size:.68rem; color:{G600}; margin-top:.12rem; }}

/* Section */
.sec {{ font-size:.63rem; font-weight:700; letter-spacing:.14em; text-transform:uppercase; color:{NY};
    margin:1.3rem 0 .5rem 0; padding-bottom:.3rem; border-bottom:2px solid {TL}; display:inline-block; }}

/* Insight callout – the "so what" above each chart (MBB style) */
.so {{ font-size:.78rem; color:{NY}; font-weight:600; margin:0 0 .6rem 0; line-height:1.4; }}

/* Summary box */
.es {{ background:{W}; border-left:4px solid {TL}; border-radius:0 4px 4px 0;
    padding:1rem 1.3rem; font-size:.8rem; line-height:1.6; color:{G800}; }}
.es b {{ color:{NY}; }}

/* Alert box */
.al {{ background:#FFF3CD; border:1px solid #FFEAA7; border-radius:4px;
    padding:.75rem 1rem; font-size:.78rem; color:#856404; margin-bottom:.6rem; }}
.al-r {{ background:#F8D7DA; border:1px solid #F5C6CB; color:#721C24; }}

/* Workflow */
.wf {{ background:{TLL}; border:1px solid {TL}33; border-radius:4px;
    padding:1rem 1.2rem; font-size:.8rem; line-height:1.5; color:{G800}; margin-bottom:.8rem; }}
.wf b {{ color:{NY}; }}
.ws {{ display:inline-flex; align-items:center; background:{NY}; color:{W};
    font-size:.62rem; font-weight:600; padding:.15rem .5rem; border-radius:3px; margin-right:.2rem; }}

/* Footer */
.ft {{ text-align:center; padding:1.5rem 0 .6rem; font-size:.58rem; color:{G400}; letter-spacing:.06em; }}

#MainMenu, footer, header {{ visibility:hidden; }}
.stDownloadButton button {{
    background:{NY} !important; color:{W} !important; border:none !important;
    font-family:'DM Sans',sans-serif !important; font-weight:600 !important; font-size:.76rem !important;
}}
</style>
""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
# CONSTANTS & HELPERS
# ═══════════════════════════════════════════════════════════════════════════════
SF_COLS   = ["Stage","Solution Resource","Account Name","Owner Role","Opportunity Name",
             "Opportunity Owner","Main Primary Service","Opportunity PAR","Stage Duration","Close Date","Notes"]
TEAM_COLS = ["Solutions Notes","Tasks","Action Items","Comments / Results"]
ALL_COLS  = SF_COLS + TEAM_COLS

STAGE_ORDER = ["Information Gathering","Solutions Design","Proposal/Price Quote","Proposal Price/Quote","Negotiations","Closed/Won","Closed/Lost"]

def clean_upload(df):
    df = df.loc[:, ~df.columns.str.startswith("Unnamed")].copy()
    rn = {}
    for c in df.columns:
        for s in SF_COLS + TEAM_COLS:
            if c.strip().lower().replace(" ","") == s.lower().replace(" ",""):
                rn[c] = s
    df.rename(columns=rn, inplace=True)
    if "Opportunity PAR" in df.columns:
        df["Opportunity PAR"] = pd.to_numeric(df["Opportunity PAR"], errors="coerce").fillna(0)
    if "Stage Duration" in df.columns:
        df["Stage Duration"] = pd.to_numeric(df["Stage Duration"], errors="coerce").fillna(0).astype(int)
    if "Close Date" in df.columns:
        df["Close Date"] = pd.to_datetime(df["Close Date"], errors="coerce", dayfirst=False).dt.strftime("%m/%d/%Y")
    return df.reset_index(drop=True)

def merge_masterfile(master, new_sf):
    for c in TEAM_COLS:
        if c not in master.columns: master[c] = ""
    old_o = set(master["Opportunity Name"].dropna())
    new_o = set(new_sf["Opportunity Name"].dropna())
    upd = 0
    for i, r in master.iterrows():
        o = r.get("Opportunity Name")
        if o in new_o:
            m = new_sf[new_sf["Opportunity Name"]==o].iloc[0]
            for c in SF_COLS:
                if c in new_sf.columns: master.at[i,c] = m[c]
            upd += 1
    added = new_o - old_o
    nr = new_sf[new_sf["Opportunity Name"].isin(added)].copy()
    for c in TEAM_COLS: nr[c] = ""
    master = pd.concat([master,nr], ignore_index=True)
    removed = old_o - new_o
    for o in removed:
        mask = master["Opportunity Name"]==o
        master.loc[mask,"Solutions Notes"] = master.loc[mask,"Solutions Notes"].fillna("").astype(str) + " [Removed from SF]"
    cols = [c for c in ALL_COLS if c in master.columns]
    return master[cols].reset_index(drop=True), {"updated":upd,"added":len(added),"removed":len(removed),"total":len(master)}

def fc(v):
    if pd.isna(v) or v==0: return "$0"
    if abs(v)>=1e6: return f"${v/1e6:,.1f}M"
    if abs(v)>=1e3: return f"${v/1e3:,.0f}K"
    return f"${v:,.0f}"

def pct(part, whole):
    return f"{part/whole*100:.0f}%" if whole else "0%"

def pl(fig, h=340, mb=30):
    fig.update_layout(
        font=dict(family="DM Sans,sans-serif", size=11, color=G800),
        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
        height=h, margin=dict(l=10,r=10,t=28,b=mb),
        legend=dict(font=dict(size=9.5), orientation="h", y=-0.2, x=0.5, xanchor="center"),
    )
    fig.update_xaxes(gridcolor=G200, showline=True, linecolor=G200, linewidth=1)
    fig.update_yaxes(gridcolor=G200, showline=True, linecolor=G200, linewidth=1)
    return fig

def to_excel(df):
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name="Masterfile")
        ws = w.sheets["Masterfile"]
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        hf=PatternFill("solid",fgColor="002B49"); hw=Font(name="Calibri",bold=True,color="FFFFFF",size=10)
        tf=PatternFill("solid",fgColor="E8F5F3"); bd=Border(*(Side(style="thin",color="E1E4E8"),)*4)
        for ci,cn in enumerate(df.columns,1):
            c=ws.cell(1,ci); c.fill,c.font,c.border=hf,hw,bd; c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
            if cn in TEAM_COLS:
                for ri in range(2,len(df)+2): ws.cell(ri,ci).fill=tf
            ml=max(len(str(cn)),*(len(str(x)) for x in df[cn].head(50))) if len(df) else len(str(cn))
            ws.column_dimensions[ws.cell(1,ci).column_letter].width=min(ml+4,42)
        bf=Font(name="Calibri",size=10)
        for ri in range(2,len(df)+2):
            for ci in range(1,len(df.columns)+1):
                c=ws.cell(ri,ci); c.font,c.border=bf,bd; c.alignment=Alignment(vertical="center",wrap_text=True)
    return buf.getvalue()

# ═══════════════════════════════════════════════════════════════════════════════
# SESSION STATE
# ═══════════════════════════════════════════════════════════════════════════════
if "master" not in st.session_state: st.session_state.master = None

# ═══════════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ═══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown(f"""
    <div style="padding:1.1rem 0 .9rem; text-align:center;">
        <div style="font-size:.56rem; font-weight:700; letter-spacing:.24em; color:{GD}; text-transform:uppercase;">Marken · UPS Healthcare</div>
        <div style="font-size:.95rem; font-weight:700; color:{W}; margin-top:.15rem;">Solutions Pipeline</div>
        <div style="font-size:.62rem; color:rgba(255,255,255,0.35); margin-top:.1rem;">Precision Logistics</div>
    </div>
    """, unsafe_allow_html=True)
    st.divider()
    page = st.radio("", ["Dashboard","Masterfile Manager"], label_visibility="collapsed")
    st.divider()
    st.markdown(f'<div style="font-size:.58rem; color:rgba(255,255,255,0.25); text-align:center; line-height:1.4;">Report generated<br>{datetime.now().strftime("%d %b %Y · %H:%M")}</div>', unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
# HEADER
# ═══════════════════════════════════════════════════════════════════════════════
st.markdown(f"""
<div class="rh">
    <div><div class="rh-b">Marken · UPS Healthcare Precision Logistics</div>
        <div class="rh-t">Solutions Team — Global Pipeline Report</div>
        <div class="rh-s">Opportunity overview · Masterfile management · Analytics</div></div>
    <div class="rh-r"><div class="rh-d">{datetime.now().strftime('%d %B %Y')}</div><div class="rh-c">Confidential</div></div>
</div>
""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
# INITIAL UPLOAD
# ═══════════════════════════════════════════════════════════════════════════════
if st.session_state.master is None:
    st.markdown("""<div class="wf"><b>Getting started</b> — Upload your Salesforce export or existing Masterfile (.xlsx / .csv). The app detects the format and adds team columns if needed.</div>""", unsafe_allow_html=True)
    f = st.file_uploader("Upload Salesforce Export or Masterfile", type=["xlsx","xls","csv"], label_visibility="collapsed")
    if f:
        raw = pd.read_csv(f) if f.name.endswith(".csv") else pd.read_excel(f)
        cl = clean_upload(raw)
        for c in TEAM_COLS:
            if c not in cl.columns: cl[c] = ""
        cols = [c for c in ALL_COLS if c in cl.columns]
        st.session_state.master = cl[cols]
        st.rerun()
    st.stop()

df = st.session_state.master.copy()
df["Opportunity PAR"] = pd.to_numeric(df.get("Opportunity PAR",0), errors="coerce").fillna(0)
df["Stage Duration"]  = pd.to_numeric(df.get("Stage Duration",0), errors="coerce").fillna(0).astype(int)
df["Close Date Parsed"] = pd.to_datetime(df["Close Date"], errors="coerce")
TODAY = pd.Timestamp.now().normalize()

# ═══════════════════════════════════════════════════════════════════════════════
#  DASHBOARD
# ═══════════════════════════════════════════════════════════════════════════════
if page == "Dashboard":

    # ── Filters ──────────────────────────────────────────────────────────────
    with st.expander("Filters", expanded=False):
        f1,f2,f3,f4 = st.columns(4)
        sel_st = f1.multiselect("Stage", sorted(df["Stage"].dropna().unique()), default=sorted(df["Stage"].dropna().unique()))
        sel_sv = f2.multiselect("Service", sorted(df["Main Primary Service"].dropna().unique()), default=sorted(df["Main Primary Service"].dropna().unique()))
        sel_rg = f3.multiselect("Region", sorted(df["Owner Role"].dropna().unique()), default=sorted(df["Owner Role"].dropna().unique()))
        sel_rs = f4.multiselect("Solution Resource", sorted(df["Solution Resource"].dropna().unique()), default=sorted(df["Solution Resource"].dropna().unique()))
    fdf = df[df["Stage"].isin(sel_st) & df["Main Primary Service"].isin(sel_sv) & df["Owner Role"].isin(sel_rg) & df["Solution Resource"].isin(sel_rs)]

    total = fdf["Opportunity PAR"].sum()
    n_opp = len(fdf)
    n_cust = fdf["Account Name"].nunique()
    n_svc = fdf["Main Primary Service"].nunique()
    avg_deal = fdf["Opportunity PAR"].mean() if n_opp else 0
    avg_dur = fdf["Stage Duration"].mean() if n_opp else 0
    past_due = fdf[fdf["Close Date Parsed"] < TODAY]
    aging_60 = fdf[fdf["Stage Duration"] > 60]

    # ── KPIs ─────────────────────────────────────────────────────────────────
    st.markdown(f"""
    <div class="kr">
        <div class="kp" style="--ac:{TL};"><div class="kp-l">Pipeline Value</div><div class="kp-v">{fc(total)}</div><div class="kp-d">{n_opp} opportunities</div></div>
        <div class="kp" style="--ac:{NY};"><div class="kp-l">Customers</div><div class="kp-v">{n_cust}</div><div class="kp-d">Unique accounts</div></div>
        <div class="kp" style="--ac:{GD};"><div class="kp-l">Products in Scope</div><div class="kp-v">{n_svc}</div><div class="kp-d">Service categories</div></div>
        <div class="kp" style="--ac:{BA};"><div class="kp-l">Avg Deal Size</div><div class="kp-v">{fc(avg_deal)}</div><div class="kp-d">Per opportunity</div></div>
        <div class="kp" style="--ac:{G400};"><div class="kp-l">Avg Stage Duration</div><div class="kp-v">{avg_dur:.0f}d</div><div class="kp-d">{len(aging_60)} over 60 days</div></div>
        <div class="kp" style="--ac:{RD};"><div class="kp-l">Past Close Date</div><div class="kp-v">{len(past_due)}</div><div class="kp-d">{fc(past_due['Opportunity PAR'].sum())} at risk</div></div>
    </div>
    """, unsafe_allow_html=True)


    # ══════════════════════════════════════════════════════════════════════════
    # 1 — PIPELINE OVERVIEW
    # ══════════════════════════════════════════════════════════════════════════
    st.markdown('<div class="sec">1 · Pipeline Overview</div>', unsafe_allow_html=True)

    p1a, p1b = st.columns([1.15, 0.85])

    with p1a:
        st.markdown('<p class="so">Solutions Design holds the bulk of the pipeline — most value is still in early stages</p>', unsafe_allow_html=True)
        sg = fdf.groupby("Stage").agg(Value=("Opportunity PAR","sum"), Count=("Opportunity Name","count")).reset_index()
        sg["StageOrd"] = sg["Stage"].apply(lambda x: STAGE_ORDER.index(x) if x in STAGE_ORDER else 99)
        sg = sg.sort_values("StageOrd")
        colors = []
        for s in sg["Stage"]:
            if "Closed/Lost" in s: colors.append(RD)
            elif "Closed/Won" in s: colors.append(GN)
            elif "Negotiation" in s: colors.append(GD)
            elif "Proposal" in s: colors.append(BA)
            elif "Design" in s: colors.append(TL)
            else: colors.append(G400)
        fig = go.Figure()
        fig.add_trace(go.Bar(
            x=sg["Stage"], y=sg["Value"], marker_color=colors,
            text=[f"{fc(v)}<br><span style='font-size:9px;color:{G600}'>{c} opps</span>" for v,c in zip(sg["Value"],sg["Count"])],
            textposition="outside", textfont=dict(size=10),
        ))
        pl(fig, h=320, mb=10)
        fig.update_layout(showlegend=False, yaxis=dict(visible=False))
        fig.update_xaxes(showgrid=False, tickfont=dict(size=9.5)); fig.update_yaxes(showgrid=False, showline=False)
        st.plotly_chart(fig, use_container_width=True)

    with p1b:
        st.markdown('<p class="so">Pipeline stage composition — value vs. count</p>', unsafe_allow_html=True)
        sg2 = sg.copy()
        sg2["Val%"] = sg2["Value"]/sg2["Value"].sum()*100
        sg2["Cnt%"] = sg2["Count"]/sg2["Count"].sum()*100
        fig2 = make_subplots(rows=1, cols=2, specs=[[{"type":"pie"},{"type":"pie"}]], subplot_titles=["By Value","By Count"])
        fig2.add_trace(go.Pie(labels=sg2["Stage"], values=sg2["Value"], hole=.5,
            marker=dict(colors=colors), textinfo="percent", textfont=dict(size=10),
            hovertemplate="<b>%{label}</b><br>$%{value:,.0f}<extra></extra>", sort=False), 1, 1)
        fig2.add_trace(go.Pie(labels=sg2["Stage"], values=sg2["Count"], hole=.5,
            marker=dict(colors=colors), textinfo="percent", textfont=dict(size=10),
            hovertemplate="<b>%{label}</b><br>%{value} opps<extra></extra>", sort=False), 1, 2)
        pl(fig2, h=320)
        fig2.update_layout(showlegend=False)
        fig2.update_annotations(font=dict(size=10, color=G600))
        st.plotly_chart(fig2, use_container_width=True)


    # ══════════════════════════════════════════════════════════════════════════
    # 2 — CUSTOMER ANALYSIS
    # ══════════════════════════════════════════════════════════════════════════
    st.markdown('<div class="sec">2 · Customer Concentration</div>', unsafe_allow_html=True)

    c2a, c2b = st.columns([1.15, 0.85])

    with c2a:
        cu = fdf.groupby("Account Name").agg(Value=("Opportunity PAR","sum"), Count=("Opportunity Name","count")).reset_index().sort_values("Value", ascending=False)
        top5_pct = cu.head(5)["Value"].sum()/cu["Value"].sum()*100 if cu["Value"].sum()>0 else 0
        st.markdown(f'<p class="so">Top 5 accounts represent {top5_pct:.0f}% of total pipeline — high concentration risk</p>', unsafe_allow_html=True)
        cu_top = cu.head(10).sort_values("Value", ascending=True)
        fig3 = go.Figure(go.Bar(
            y=cu_top["Account Name"], x=cu_top["Value"], orientation="h",
            marker=dict(color=cu_top["Value"], colorscale=[[0,"#B2DFDB"],[.5,TL],[1,NY]], showscale=False),
            text=[f"  {fc(v)}  ({c})" for v,c in zip(cu_top["Value"],cu_top["Count"])],
            textposition="outside", textfont=dict(size=10, color=G800),
        ))
        pl(fig3, h=max(280, 34*len(cu_top)))
        fig3.update_layout(showlegend=False, xaxis=dict(visible=False), yaxis=dict(tickfont=dict(size=9.5)))
        fig3.update_xaxes(showgrid=False, showline=False); fig3.update_yaxes(showgrid=False, showline=False)
        st.plotly_chart(fig3, use_container_width=True)

    with c2b:
        st.markdown('<p class="so">Cumulative concentration (Pareto curve)</p>', unsafe_allow_html=True)
        cu_sorted = cu.sort_values("Value", ascending=False).reset_index(drop=True)
        cu_sorted["CumVal"] = cu_sorted["Value"].cumsum()
        cu_sorted["CumPct"] = cu_sorted["CumVal"]/cu_sorted["Value"].sum()*100
        cu_sorted["Rank"] = range(1, len(cu_sorted)+1)
        fig4 = go.Figure()
        fig4.add_trace(go.Bar(x=cu_sorted["Rank"], y=cu_sorted["Value"], marker_color=TL, name="Individual Value",
            hovertemplate="<b>%{customdata}</b><br>$%{y:,.0f}<extra></extra>", customdata=cu_sorted["Account Name"]))
        fig4.add_trace(go.Scatter(x=cu_sorted["Rank"], y=cu_sorted["CumPct"], mode="lines+markers",
            line=dict(color=NY, width=2.5), marker=dict(size=5, color=NY), name="Cumulative %", yaxis="y2"))
        fig4.add_hline(y=80, line_dash="dot", line_color=RD, opacity=0.5, annotation_text="80%",
            annotation_position="right", yref="y2")
        pl(fig4, h=max(280, 34*len(cu_top)), mb=40)
        fig4.update_layout(
            yaxis=dict(title="Value ($)", titlefont=dict(size=9), visible=False),
            yaxis2=dict(title="Cumulative %", titlefont=dict(size=9), side="right", overlaying="y", range=[0,105], showgrid=False),
            xaxis=dict(title="Customer Rank", titlefont=dict(size=9), tickfont=dict(size=9)),
            legend=dict(font=dict(size=8)),
        )
        fig4.update_xaxes(showgrid=False); fig4.update_yaxes(showgrid=False, showline=False)
        st.plotly_chart(fig4, use_container_width=True)


    # ══════════════════════════════════════════════════════════════════════════
    # 3 — PRODUCT / SERVICE MIX
    # ══════════════════════════════════════════════════════════════════════════
    st.markdown('<div class="sec">3 · Product / Service Mix</div>', unsafe_allow_html=True)

    s3a, s3b = st.columns([0.55, 0.45])

    with s3a:
        sv = fdf.groupby("Main Primary Service").agg(Value=("Opportunity PAR","sum"), Count=("Opportunity Name","count"),
            Avg=("Opportunity PAR","mean")).reset_index().sort_values("Value", ascending=False)
        top_svc = sv.iloc[0]["Main Primary Service"] if len(sv) else "N/A"
        top_svc_pct = sv.iloc[0]["Value"]/sv["Value"].sum()*100 if len(sv) and sv["Value"].sum()>0 else 0
        st.markdown(f'<p class="so">{top_svc} accounts for {top_svc_pct:.0f}% of pipeline value by service</p>', unsafe_allow_html=True)
        fig5 = go.Figure(go.Pie(
            labels=sv["Main Primary Service"], values=sv["Value"], hole=.5,
            marker=dict(colors=SEQ[:len(sv)]),
            textinfo="label+percent", textfont=dict(size=10),
            hovertemplate="<b>%{label}</b><br>$%{value:,.0f}<br>%{percent}<extra></extra>", sort=True,
        ))
        pl(fig5, h=310); fig5.update_layout(showlegend=False)
        st.plotly_chart(fig5, use_container_width=True)

    with s3b:
        st.markdown('<p class="so">Average deal size varies significantly across service types</p>', unsafe_allow_html=True)
        sv_s = sv.sort_values("Avg", ascending=True)
        fig6 = go.Figure(go.Bar(
            y=sv_s["Main Primary Service"], x=sv_s["Avg"], orientation="h", marker_color=NY,
            text=[f"  {fc(v)}  ({c} opps)" for v,c in zip(sv_s["Avg"],sv_s["Count"])],
            textposition="outside", textfont=dict(size=9.5, color=G800),
        ))
        pl(fig6, h=310)
        fig6.update_layout(showlegend=False, xaxis=dict(visible=False), yaxis=dict(tickfont=dict(size=9.5)))
        fig6.update_xaxes(showgrid=False, showline=False); fig6.update_yaxes(showgrid=False, showline=False)
        st.plotly_chart(fig6, use_container_width=True)

    # Service × Region Heatmap
    st.markdown('<p class="so">Service demand mapped by region — Next Flight Out dominates both EMEA and NORAM</p>', unsafe_allow_html=True)
    ht = pd.crosstab(fdf["Main Primary Service"], fdf["Owner Role"], values=fdf["Opportunity PAR"], aggfunc="sum").fillna(0)
    fig7 = go.Figure(go.Heatmap(
        z=ht.values, x=ht.columns.tolist(), y=ht.index.tolist(),
        colorscale=[[0,W],[.3,TLL],[.7,TL],[1,NY]],
        text=[[fc(v) for v in row] for row in ht.values], texttemplate="%{text}", textfont=dict(size=10),
        hovertemplate="Service: %{y}<br>Region: %{x}<br>Value: $%{z:,.0f}<extra></extra>",
        showscale=False,
    ))
    pl(fig7, h=max(200, 36*len(ht)))
    fig7.update_layout(xaxis=dict(tickfont=dict(size=10), side="top"), yaxis=dict(tickfont=dict(size=10), autorange="reversed"))
    fig7.update_xaxes(showgrid=False, showline=False); fig7.update_yaxes(showgrid=False, showline=False)
    st.plotly_chart(fig7, use_container_width=True)


    # ══════════════════════════════════════════════════════════════════════════
    # 4 — REGIONAL ANALYSIS
    # ══════════════════════════════════════════════════════════════════════════
    st.markdown('<div class="sec">4 · Regional Split</div>', unsafe_allow_html=True)

    r4a, r4b = st.columns(2)

    with r4a:
        rg = fdf.groupby("Owner Role").agg(Value=("Opportunity PAR","sum"), Count=("Opportunity Name","count"),
            Avg=("Opportunity PAR","mean"), AvgDur=("Stage Duration","mean")).reset_index().sort_values("Value", ascending=False)
        top_reg = rg.iloc[0]["Owner Role"] if len(rg) else "N/A"
        st.markdown(f'<p class="so">{top_reg} leads the pipeline by value</p>', unsafe_allow_html=True)
        fig8 = go.Figure(go.Bar(
            x=rg["Owner Role"], y=rg["Value"],
            marker_color=[NY,TL,GD,BA,G600][:len(rg)],
            text=[f"{fc(v)}<br><span style='font-size:9px'>{c} opps · Avg {fc(a)}</span>" for v,c,a in zip(rg["Value"],rg["Count"],rg["Avg"])],
            textposition="inside", textfont=dict(size=11, color=W),
        ))
        pl(fig8, h=300)
        fig8.update_layout(showlegend=False, yaxis=dict(visible=False))
        fig8.update_xaxes(showgrid=False, tickfont=dict(size=10)); fig8.update_yaxes(showgrid=False, showline=False)
        st.plotly_chart(fig8, use_container_width=True)

    with r4b:
        st.markdown('<p class="so">Region × stage: which regions are further in the pipeline?</p>', unsafe_allow_html=True)
        rs_ht = pd.crosstab(fdf["Owner Role"], fdf["Stage"], values=fdf["Opportunity PAR"], aggfunc="sum").fillna(0)
        ordered = [s for s in STAGE_ORDER if s in rs_ht.columns]
        extra = [s for s in rs_ht.columns if s not in STAGE_ORDER]
        rs_ht = rs_ht[ordered + extra]
        stg_colors = {
            "Information Gathering": G400, "Solutions Design": TL, "Proposal/Price Quote": BA,
            "Proposal Price/Quote": BA, "Negotiations": GD, "Closed/Won": GN, "Closed/Lost": RD,
        }
        fig9 = go.Figure()
        for col in rs_ht.columns:
            fig9.add_trace(go.Bar(
                x=rs_ht.index, y=rs_ht[col], name=col, marker_color=stg_colors.get(col, G400),
                text=[fc(v) if v > 0 else "" for v in rs_ht[col]], textposition="inside", textfont=dict(size=9, color=W),
            ))
        pl(fig9, h=300, mb=10)
        fig9.update_layout(barmode="stack", yaxis=dict(visible=False), legend=dict(font=dict(size=8)))
        fig9.update_xaxes(showgrid=False, tickfont=dict(size=10)); fig9.update_yaxes(showgrid=False, showline=False)
        st.plotly_chart(fig9, use_container_width=True)


    # ══════════════════════════════════════════════════════════════════════════
    # 5 — SOLUTION RESOURCE WORKLOAD
    # ══════════════════════════════════════════════════════════════════════════
    st.markdown('<div class="sec">5 · Solution Resource Workload</div>', unsafe_allow_html=True)

    rw = fdf.groupby("Solution Resource").agg(Value=("Opportunity PAR","sum"), Count=("Opportunity Name","count"),
        AvgDur=("Stage Duration","mean"), Cust=("Account Name","nunique")).reset_index().sort_values("Value", ascending=False)

    r5a, r5b = st.columns([0.6, 0.4])

    with r5a:
        st.markdown(f'<p class="so">{rw.iloc[0]["Solution Resource"]} carries the largest pipeline at {fc(rw.iloc[0]["Value"])}</p>' if len(rw) else '', unsafe_allow_html=True)
        fig10 = go.Figure()
        fig10.add_trace(go.Bar(x=rw["Solution Resource"], y=rw["Count"], name="# Opportunities", marker_color=NY,
            text=rw["Count"], textposition="auto", textfont=dict(color=W, size=10)))
        fig10.add_trace(go.Scatter(x=rw["Solution Resource"], y=rw["Value"], name="Total Value ($)",
            mode="markers+lines", marker=dict(color=GD, size=9, line=dict(width=1.5, color=NY)),
            line=dict(color=GD, width=2), yaxis="y2"))
        pl(fig10, h=310, mb=50)
        fig10.update_layout(
            yaxis=dict(title="# Opps", titlefont=dict(size=9), side="left"),
            yaxis2=dict(title="Value ($)", titlefont=dict(size=9), side="right", overlaying="y", showgrid=False),
            xaxis=dict(tickangle=15, tickfont=dict(size=9)),
        )
        st.plotly_chart(fig10, use_container_width=True)

    with r5b:
        st.markdown('<p class="so">Resource detail breakdown</p>', unsafe_allow_html=True)
        rd = rw.rename(columns={"Solution Resource":"Resource","Count":"Opps","Value":"Pipeline","AvgDur":"Avg Days","Cust":"Customers"}).copy()
        rd["Pipeline"] = rd["Pipeline"].apply(lambda x: f"${x:,.0f}")
        rd["Avg Days"] = rd["Avg Days"].apply(lambda x: f"{x:.0f}")
        st.dataframe(rd, use_container_width=True, height=310, hide_index=True)

    # Resource × Region
    st.markdown('<p class="so">Resource allocation by region</p>', unsafe_allow_html=True)
    rr = pd.crosstab(fdf["Solution Resource"], fdf["Owner Role"], values=fdf["Opportunity Name"], aggfunc="count").fillna(0).astype(int)
    fig11 = go.Figure(go.Heatmap(
        z=rr.values, x=rr.columns.tolist(), y=rr.index.tolist(),
        colorscale=[[0,W],[.5,TLL],[1,NY]],
        text=rr.values, texttemplate="%{text}", textfont=dict(size=11),
        showscale=False,
    ))
    pl(fig11, h=max(180, 32*len(rr)))
    fig11.update_layout(xaxis=dict(tickfont=dict(size=10), side="top"), yaxis=dict(tickfont=dict(size=10), autorange="reversed"))
    fig11.update_xaxes(showgrid=False, showline=False); fig11.update_yaxes(showgrid=False, showline=False)
    st.plotly_chart(fig11, use_container_width=True)


    # ══════════════════════════════════════════════════════════════════════════
    # 6 — TIMELINE & AGING
    # ══════════════════════════════════════════════════════════════════════════
    st.markdown('<div class="sec">6 · Timeline & Aging Analysis</div>', unsafe_allow_html=True)

    t6a, t6b = st.columns(2)

    with t6a:
        st.markdown('<p class="so">Expected revenue by close month</p>', unsafe_allow_html=True)
        tl = fdf.dropna(subset=["Close Date Parsed"]).copy()
        tl["Month"] = tl["Close Date Parsed"].dt.to_period("M").dt.to_timestamp()
        mo = tl.groupby("Month").agg(Value=("Opportunity PAR","sum"), Count=("Opportunity Name","count")).reset_index().sort_values("Month")
        fig12 = go.Figure()
        fig12.add_trace(go.Bar(x=mo["Month"], y=mo["Value"], marker_color=TL, name="Pipeline Value",
            text=[f"{fc(v)}" for v in mo["Value"]], textposition="outside", textfont=dict(size=9)))
        fig12.add_trace(go.Scatter(x=mo["Month"], y=mo["Count"], mode="markers+lines",
            marker=dict(color=NY, size=7), line=dict(color=NY, width=2), name="# Opps", yaxis="y2"))
        pl(fig12, h=310, mb=40)
        fig12.update_layout(
            yaxis=dict(visible=False),
            yaxis2=dict(title="# Opps", titlefont=dict(size=9), side="right", overlaying="y", showgrid=False),
            xaxis=dict(tickformat="%b '%y", tickfont=dict(size=9)),
        )
        fig12.update_xaxes(showgrid=False); fig12.update_yaxes(showgrid=False, showline=False)
        st.plotly_chart(fig12, use_container_width=True)

    with t6b:
        st.markdown('<p class="so">Stage duration distribution — half the pipeline is aging over 60 days</p>', unsafe_allow_html=True)
        stages_for_box = [s for s in STAGE_ORDER if s in fdf["Stage"].unique()]
        fig13 = go.Figure()
        for s in stages_for_box:
            sd = fdf[fdf["Stage"]==s]
            fig13.add_trace(go.Box(y=sd["Stage Duration"], name=s, marker_color=TL, boxmean=True,
                fillcolor=TLL, line=dict(color=TL)))
        pl(fig13, h=310)
        fig13.update_layout(showlegend=False, yaxis=dict(title="Days", titlefont=dict(size=9)), xaxis=dict(tickfont=dict(size=9)))
        st.plotly_chart(fig13, use_container_width=True)

    # Bubble chart: value vs duration vs close date
    st.markdown('<p class="so">Opportunity landscape: value vs. stage duration (bubble = deal size, color = stage)</p>', unsafe_allow_html=True)
    bdf = fdf.dropna(subset=["Close Date Parsed"]).copy()
    stg_c = {"Information Gathering": G400, "Solutions Design": TL, "Proposal/Price Quote": BA,
             "Proposal Price/Quote": BA, "Negotiations": GD, "Closed/Won": GN, "Closed/Lost": RD}
    fig14 = go.Figure()
    for stg in bdf["Stage"].unique():
        sd = bdf[bdf["Stage"]==stg]
        fig14.add_trace(go.Scatter(
            x=sd["Stage Duration"], y=sd["Opportunity PAR"], mode="markers", name=stg,
            marker=dict(size=sd["Opportunity PAR"].apply(lambda x: max(7, min(35, x/80000))),
                        color=stg_c.get(stg, G400), line=dict(width=1, color=W), opacity=0.85),
            text=sd["Account Name"],
            hovertemplate="<b>%{text}</b><br>Duration: %{x}d<br>Value: $%{y:,.0f}<extra></extra>",
        ))
    fig14.add_vline(x=60, line_dash="dot", line_color=RD, opacity=0.4, annotation_text="60d", annotation_position="top")
    pl(fig14, h=340, mb=40)
    fig14.update_layout(
        xaxis=dict(title="Stage Duration (days)", titlefont=dict(size=10)),
        yaxis=dict(title="PAR Value ($)", titlefont=dict(size=10)),
        legend=dict(font=dict(size=8.5)),
    )
    st.plotly_chart(fig14, use_container_width=True)

    # Aging table
    st.markdown('<p class="so">Aging flags — opportunities requiring attention</p>', unsafe_allow_html=True)
    ag = fdf.copy()
    ag["Flag"] = ""
    ag.loc[ag["Stage Duration"] > 90, "Flag"] = "🔴 >90d"
    ag.loc[(ag["Stage Duration"] > 60) & (ag["Stage Duration"] <= 90), "Flag"] = "🟡 >60d"
    ag.loc[ag["Close Date Parsed"] < TODAY, "Flag"] = ag.loc[ag["Close Date Parsed"] < TODAY, "Flag"].astype(str) + " ⚠ Past Due"
    ag_flagged = ag[ag["Flag"].str.len() > 0].sort_values("Stage Duration", ascending=False)
    if len(ag_flagged):
        st.dataframe(
            ag_flagged[["Flag","Account Name","Opportunity Name","Stage","Opportunity PAR","Stage Duration","Close Date","Notes"]].style.format({"Opportunity PAR":"${:,.0f}"}),
            use_container_width=True, height=min(300, 35*len(ag_flagged)+38), hide_index=True,
        )
    else:
        st.success("No aging flags — all opportunities within normal parameters.")


    # ══════════════════════════════════════════════════════════════════════════
    # 7 — RISK & CLOSED/LOST
    # ══════════════════════════════════════════════════════════════════════════
    st.markdown('<div class="sec">7 · Risk & Attention Items</div>', unsafe_allow_html=True)

    r7a, r7b = st.columns(2)

    with r7a:
        lost = fdf[fdf["Stage"]=="Closed/Lost"]
        if len(lost):
            st.markdown(f'<div class="al al-r"><b>{len(lost)} Closed/Lost</b> opportunities totaling <b>{fc(lost["Opportunity PAR"].sum())}</b></div>', unsafe_allow_html=True)
            for _, r in lost.iterrows():
                st.markdown(f"""
                <div style="background:{W}; border:1px solid {G200}; border-radius:4px; padding:.7rem .9rem; margin-bottom:.4rem; font-size:.78rem;">
                    <b style="color:{NY};">{r['Account Name']}</b> · {fc(r['Opportunity PAR'])}<br>
                    <span style="color:{G600};">{r['Notes'] if pd.notna(r['Notes']) else 'No notes'}</span>
                </div>
                """, unsafe_allow_html=True)
        else:
            st.success("No Closed/Lost opportunities in current view.")

    with r7b:
        st.markdown('<p class="so">Largest deals at risk (high value + high duration)</p>', unsafe_allow_html=True)
        risk = fdf[~fdf["Stage"].str.contains("Closed", na=False)].copy()
        risk["RiskScore"] = risk["Opportunity PAR"] * np.log1p(risk["Stage Duration"])
        risk_top = risk.nlargest(5, "RiskScore")
        if len(risk_top):
            fig15 = go.Figure(go.Bar(
                y=risk_top["Account Name"], x=risk_top["Opportunity PAR"], orientation="h",
                marker_color=[RD if d>60 else GD if d>30 else TL for d in risk_top["Stage Duration"]],
                text=[f"  {fc(v)} · {d}d" for v,d in zip(risk_top["Opportunity PAR"],risk_top["Stage Duration"])],
                textposition="outside", textfont=dict(size=10, color=G800),
            ))
            pl(fig15, h=220)
            fig15.update_layout(showlegend=False, xaxis=dict(visible=False), yaxis=dict(tickfont=dict(size=9.5)))
            fig15.update_xaxes(showgrid=False, showline=False); fig15.update_yaxes(showgrid=False, showline=False)
            st.plotly_chart(fig15, use_container_width=True)


    # ══════════════════════════════════════════════════════════════════════════
    # 8 — OPPORTUNITY OWNER ANALYSIS
    # ══════════════════════════════════════════════════════════════════════════
    st.markdown('<div class="sec">8 · Opportunity Owner Performance</div>', unsafe_allow_html=True)

    ow = fdf.groupby("Opportunity Owner").agg(Value=("Opportunity PAR","sum"), Count=("Opportunity Name","count"),
        Avg=("Opportunity PAR","mean"), AvgDur=("Stage Duration","mean")).reset_index().sort_values("Value", ascending=False)

    o8a, o8b = st.columns([0.55, 0.45])
    with o8a:
        st.markdown('<p class="so">Top opportunity owners by pipeline value</p>', unsafe_allow_html=True)
        ow_top = ow.head(10).sort_values("Value", ascending=True)
        fig16 = go.Figure(go.Bar(
            y=ow_top["Opportunity Owner"], x=ow_top["Value"], orientation="h", marker_color=NY,
            text=[f"  {fc(v)} ({c})" for v,c in zip(ow_top["Value"],ow_top["Count"])],
            textposition="outside", textfont=dict(size=9.5, color=G800),
        ))
        pl(fig16, h=max(280, 32*len(ow_top)))
        fig16.update_layout(showlegend=False, xaxis=dict(visible=False), yaxis=dict(tickfont=dict(size=9.5)))
        fig16.update_xaxes(showgrid=False, showline=False); fig16.update_yaxes(showgrid=False, showline=False)
        st.plotly_chart(fig16, use_container_width=True)
    with o8b:
        st.markdown('<p class="so">Owner performance table</p>', unsafe_allow_html=True)
        od = ow.rename(columns={"Opportunity Owner":"Owner","Count":"Opps","Value":"Pipeline","Avg":"Avg Deal","AvgDur":"Avg Days"}).copy()
        od["Pipeline"] = od["Pipeline"].apply(lambda x: f"${x:,.0f}")
        od["Avg Deal"] = od["Avg Deal"].apply(lambda x: f"${x:,.0f}")
        od["Avg Days"] = od["Avg Days"].apply(lambda x: f"{x:.0f}")
        st.dataframe(od, use_container_width=True, height=max(280, 32*len(ow_top)), hide_index=True)


    # ══════════════════════════════════════════════════════════════════════════
    # 9 — FULL PIPELINE TABLE
    # ══════════════════════════════════════════════════════════════════════════
    st.markdown('<div class="sec">9 · Full Pipeline Detail</div>', unsafe_allow_html=True)
    tbl = fdf[["Stage","Account Name","Opportunity Name","Solution Resource","Opportunity Owner",
               "Main Primary Service","Opportunity PAR","Stage Duration","Close Date","Notes"]].copy()
    tbl = tbl.sort_values(["Stage","Opportunity PAR"], ascending=[True,False])
    st.dataframe(
        tbl.style.format({"Opportunity PAR":"${:,.0f}"}),
        use_container_width=True, height=min(500, 35*len(tbl)+38), hide_index=True,
        column_config={
            "Account Name": st.column_config.TextColumn("Customer", width="medium"),
            "Opportunity Name": st.column_config.TextColumn("Opportunity", width="large"),
            "Main Primary Service": st.column_config.TextColumn("Service", width="medium"),
        },
    )


    # ══════════════════════════════════════════════════════════════════════════
    # 10 — EXECUTIVE SUMMARY
    # ══════════════════════════════════════════════════════════════════════════
    st.markdown('<div class="sec">10 · Executive Summary</div>', unsafe_allow_html=True)

    top_c = fdf.groupby("Account Name")["Opportunity PAR"].sum().idxmax() if n_opp else "N/A"
    top_cv = fdf.groupby("Account Name")["Opportunity PAR"].sum().max() if n_opp else 0
    top_sv = fdf.groupby("Main Primary Service")["Opportunity PAR"].sum().idxmax() if n_opp else "N/A"
    n_des = len(fdf[fdf["Stage"]=="Solutions Design"])
    n_pro = len(fdf[fdf["Stage"].str.contains("Proposal|Price", case=False, na=False)])
    n_neg = len(fdf[fdf["Stage"]=="Negotiations"])
    top5p = cu.head(5)["Value"].sum()/cu["Value"].sum()*100 if len(cu) and cu["Value"].sum()>0 else 0

    reg_lead = f"with <b>{top_reg}</b> leading at <b>{fc(rg.iloc[0]['Value'])}</b>" if len(rg) else ""

    st.markdown(f"""<div class="es">
The Solutions team is currently managing <b>{n_opp} active opportunities</b> representing
a total pipeline value of <b>{fc(total)}</b> across <b>{n_cust} unique customers</b>
and <b>{n_svc} service categories</b>.
<br><br>
<b>Stage composition:</b> {n_des} opportunities are in Solutions Design ({pct(n_des,n_opp)}),
{n_pro} in Proposal/Price Quote, {n_neg} in Negotiations, and {len(lost)} were Closed/Lost
({fc(lost['Opportunity PAR'].sum())} lost value).
<br><br>
<b>Customer concentration:</b> The top 5 accounts represent <b>{top5p:.0f}%</b> of pipeline value,
led by <b>{top_c}</b> at {fc(top_cv)}. This is a notable concentration risk.
The dominant service type is <b>{top_sv}</b>.
<br><br>
<b>Regional coverage:</b> Pipeline spans <b>{fdf['Owner Role'].nunique()} regions</b>
{reg_lead}.
<b>{fdf['Solution Resource'].nunique()} Solution Resources</b> are actively engaged.
<br><br>
<b>Attention items:</b> <b>{len(past_due)} opportunities</b> ({fc(past_due['Opportunity PAR'].sum())})
have close dates that have already passed.
<b>{len(aging_60)} opportunities</b> have been in their current stage for over 60 days.
Average stage duration stands at <b>{avg_dur:.0f} days</b>.
</div>""", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════════
#  MASTERFILE MANAGER
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "Masterfile Manager":

    st.markdown(f"""
    <div class="wf">
        <b>Workflow</b><br><br>
        <span class="ws">1</span> Download fresh data from Salesforce<br>
        <span class="ws">2</span> Upload here — the app <b>merges</b> new SF data into the Masterfile<br>
        <span class="ws">3</span> Salesforce fields refresh · <b>Team columns preserved</b> (Solutions Notes, Tasks, Action Items, Comments)<br>
        <span class="ws">4</span> Edit team columns inline below<br>
        <span class="ws">5</span> Download the updated Masterfile
    </div>
    """, unsafe_allow_html=True)

    st.markdown('<div class="sec">Upload New Salesforce Export to Merge</div>', unsafe_allow_html=True)
    mf = st.file_uploader("Upload new SF export. Team columns will be preserved.", type=["xlsx","xls","csv"], key="mu")
    if mf:
        raw = pd.read_csv(mf) if mf.name.endswith(".csv") else pd.read_excel(mf)
        new_sf = clean_upload(raw)
        merged, stats = merge_masterfile(st.session_state.master.copy(), new_sf)
        st.session_state.master = merged
        st.success(f"Merge complete — **{stats['updated']}** updated · **{stats['added']}** added · **{stats['removed']}** flagged · **{stats['total']}** total")
        st.rerun()

    st.markdown('<div class="sec">Masterfile — Editable</div>', unsafe_allow_html=True)
    st.caption("Salesforce columns are locked. Edit the four team columns (teal-highlighted in Excel download).")

    edf = st.session_state.master.copy()
    for c in TEAM_COLS:
        if c not in edf.columns: edf[c] = ""

    edited = st.data_editor(
        edf, use_container_width=True, height=min(600, 35*len(edf)+38), num_rows="dynamic",
        column_config={
            "Stage": st.column_config.TextColumn("Stage", disabled=True),
            "Account Name": st.column_config.TextColumn("Customer", disabled=True),
            "Opportunity Name": st.column_config.TextColumn("Opportunity", disabled=True, width="large"),
            "Opportunity PAR": st.column_config.NumberColumn("PAR ($)", format="$%d", disabled=True),
            "Stage Duration": st.column_config.NumberColumn("Days", disabled=True),
            "Close Date": st.column_config.TextColumn("Close", disabled=True),
            "Owner Role": st.column_config.TextColumn("Region", disabled=True),
            "Opportunity Owner": st.column_config.TextColumn("Opp Owner", disabled=True),
            "Solution Resource": st.column_config.TextColumn("Sol. Resource", disabled=True),
            "Main Primary Service": st.column_config.TextColumn("Service", disabled=True),
            "Notes": st.column_config.TextColumn("SF Notes", disabled=True),
            "Solutions Notes": st.column_config.TextColumn("Solutions Notes", width="large"),
            "Tasks": st.column_config.TextColumn("Tasks", width="large"),
            "Action Items": st.column_config.TextColumn("Action Items", width="large"),
            "Comments / Results": st.column_config.TextColumn("Comments / Results", width="large"),
        },
        hide_index=True, key="editor",
    )

    if st.button("Save edits", type="primary"):
        st.session_state.master = edited.copy()
        st.success("Edits saved to session.")

    st.markdown("---")
    d1, d2, _ = st.columns([1,1,2])
    with d1:
        st.download_button("Download Masterfile (.xlsx)", data=to_excel(st.session_state.master),
            file_name=f"Solutions_Masterfile_{datetime.now().strftime('%Y-%m-%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    with d2:
        st.download_button("Download Masterfile (.csv)", data=st.session_state.master.to_csv(index=False).encode(),
            file_name=f"Solutions_Masterfile_{datetime.now().strftime('%Y-%m-%d')}.csv", mime="text/csv")

# Footer
st.markdown(f'<div class="ft">MARKEN · UPS HEALTHCARE PRECISION LOGISTICS &nbsp;|&nbsp; SOLUTIONS TEAM PIPELINE REPORT &nbsp;|&nbsp; CONFIDENTIAL</div>', unsafe_allow_html=True)
