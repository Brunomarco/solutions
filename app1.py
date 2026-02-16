import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import numpy as np
import re
from io import BytesIO
from datetime import datetime

# ═══════════════════════════════════════════════════════════════════════════════
# CONFIG
# ═══════════════════════════════════════════════════════════════════════════════
st.set_page_config(page_title="Solutions Pipeline | Marken", page_icon="📋", layout="wide", initial_sidebar_state="expanded")

# ── Marken Brand Palette ─────────────────────────────────────────────────────
NY    = "#002B49"
NY2   = "#003A63"
TL    = "#00857C"
TLL   = "#E8F5F3"
GD    = "#FFB500"
W     = "#FFFFFF"
G50   = "#FAFBFC"
G100  = "#F4F5F7"
G200  = "#E1E4E8"
G400  = "#A0A8B4"
G600  = "#6B7280"
G800  = "#2D3748"
BA    = "#2E86AB"
RD    = "#DC3545"
GN    = "#28A745"
OR    = "#E06C47"
AM    = "#F6AD55"
SEQ   = [NY, TL, GD, BA, "#6C5B7B", OR, GN, "#9B59B6", "#F39C12", "#1ABC9C"]

# Date display format
DATE_FMT = "%d-%b-%Y"   # e.g. 09-Feb-2026

# ── CSS ──────────────────────────────────────────────────────────────────────
st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;600;700&family=Instrument+Serif&display=swap');
.stApp {{ background:{G50}; font-family:'DM Sans',sans-serif; color:{G800}; }}
h1,h2,h3,h4 {{ font-family:'DM Sans',sans-serif; color:{NY}; }}

/* Sidebar */
section[data-testid="stSidebar"] {{ background:{NY}; }}
section[data-testid="stSidebar"] * {{ color:{W} !important; font-family:'DM Sans',sans-serif; }}
section[data-testid="stSidebar"] hr {{ border-color:rgba(255,255,255,0.10); }}

/* Header */
.rh {{ background:linear-gradient(135deg, {NY} 0%, {NY2} 100%); padding:1.8rem 2.2rem 1.6rem;
    border-radius:8px; margin-bottom:1.6rem;
    display:flex; justify-content:space-between; align-items:flex-end; }}
.rh-b {{ font-size:.58rem; font-weight:600; letter-spacing:.24em; text-transform:uppercase; color:{GD}; margin-bottom:.35rem; }}
.rh-t {{ font-family:'Instrument Serif',Georgia,serif; font-size:1.55rem; color:{W}; line-height:1.25; }}
.rh-s {{ font-size:.74rem; color:rgba(255,255,255,0.45); margin-top:.25rem; }}
.rh-r {{ text-align:right; }}
.rh-d {{ font-size:.7rem; color:rgba(255,255,255,0.40); }}
.rh-c {{ font-size:.54rem; color:{GD}; font-weight:600; letter-spacing:.14em; text-transform:uppercase; margin-top:.15rem; }}

/* KPIs */
.kr {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(155px,1fr)); gap:1rem; margin-bottom:2.2rem; }}
.kp {{ background:{W}; border:1px solid {G200}; border-top:3px solid var(--ac,{TL}); border-radius:6px;
    padding:1.1rem 1.2rem 1rem; transition:box-shadow 0.2s ease; }}
.kp:hover {{ box-shadow:0 3px 12px rgba(0,43,73,0.07); }}
.kp-l {{ font-size:.56rem; font-weight:600; letter-spacing:.12em; text-transform:uppercase; color:{G400}; margin-bottom:.3rem; }}
.kp-v {{ font-size:1.35rem; font-weight:700; color:{NY}; letter-spacing:-.03em; line-height:1.15; }}
.kp-d {{ font-size:.64rem; color:{G600}; margin-top:.15rem; }}

/* Section header — generous spacing */
.sec {{ font-size:.62rem; font-weight:700; letter-spacing:.16em; text-transform:uppercase; color:{NY};
    margin:2.8rem 0 .8rem 0; padding-bottom:.4rem; border-bottom:2px solid {TL}; display:inline-block; }}

/* MBB insight callout */
.so {{ font-size:.78rem; color:{NY}; font-weight:600; margin:.3rem 0 1rem 0; line-height:1.45; }}

/* Card wrapper for chart groups */
.cw {{ background:{W}; border:1px solid {G200}; border-radius:6px; padding:1.3rem 1.4rem 1rem; margin-bottom:.5rem; }}

/* Summary box */
.es {{ background:{W}; border-left:4px solid {TL}; border-radius:0 6px 6px 0;
    padding:1.3rem 1.6rem; font-size:.82rem; line-height:1.75; color:{G800}; }}
.es b {{ color:{NY}; }}

/* Alert box */
.al {{ background:#FFF3CD; border:1px solid #FFEAA7; border-radius:5px;
    padding:.85rem 1.1rem; font-size:.78rem; color:#856404; margin-bottom:.7rem; }}
.al-r {{ background:#F8D7DA; border:1px solid #F5C6CB; color:#721C24; }}

/* Workflow */
.wf {{ background:{TLL}; border:1px solid {TL}33; border-radius:6px;
    padding:1.1rem 1.4rem; font-size:.8rem; line-height:1.6; color:{G800}; margin-bottom:1rem; }}
.wf b {{ color:{NY}; }}
.ws {{ display:inline-flex; align-items:center; background:{NY}; color:{W};
    font-size:.6rem; font-weight:600; padding:.18rem .55rem; border-radius:3px; margin-right:.25rem; }}

/* Spacer */
.sp {{ height:1rem; }}
.sp-lg {{ height:1.8rem; }}

/* Footer */
.ft {{ text-align:center; padding:2rem 0 .8rem; font-size:.56rem; color:{G400}; letter-spacing:.08em; }}

#MainMenu, footer, header {{ visibility:hidden; }}
.stDownloadButton button {{
    background:{NY} !important; color:{W} !important; border:none !important;
    font-family:'DM Sans',sans-serif !important; font-weight:600 !important; font-size:.76rem !important;
    border-radius:5px !important;
}}
</style>
""", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════════
# CONSTANTS & HELPERS
# ═══════════════════════════════════════════════════════════════════════════════
SF_COLS   = ["Stage","Solution Resource","Account Name","Owner Role","Opportunity Name",
             "Opportunity Owner","Main Primary Service","Opportunity PAR","Stage Duration",
             "Close Date","Notes","Status","Received by Solutions","Closed by Solutions","Product"]
TEAM_COLS = ["Solutions Notes","Tasks","Action Items","Comments / Results"]
ALL_COLS  = SF_COLS + TEAM_COLS

STAGE_ORDER = ["Information Gathering","Solutions Design","Proposal/Price Quote",
               "Proposal Price/Quote","Negotiations","Closed/Won","Closed/Lost"]

STATUS_COLORS = {"Working": GN, "Pending": AM, "Completed": BA, "Unassigned": G400}

# ── Spacer helper ────────────────────────────────────────────────────────────
def spacer(size="md"):
    """Insert vertical whitespace between sections."""
    h = {"sm": "0.8rem", "md": "1.6rem", "lg": "2.8rem", "xl": "3.5rem"}.get(size, "1.6rem")
    st.markdown(f'<div style="height:{h}"></div>', unsafe_allow_html=True)


def parse_par(val):
    if pd.isna(val): return 0.0
    if isinstance(val, (int, float)): return float(val)
    s = str(val).strip().upper().replace("USD","").replace("$","").replace(",","").strip()
    try: return float(s)
    except ValueError: return 0.0


def fix_excel_eu_date(val):
    """Robust parser for Solutions date columns with mixed formats.
    Handles: DD/M/YYYY strings, M/D/YY strings (2-digit year),
    and Excel datetimes where month↔day was swapped on import.
    Examples:
        '27/1/2026'  → 27-Jan-2026   (DD/M/YYYY string)
        '2/10/26'    → 10-Feb-2026   (M/DD/YY string)
        '2/9/26'     → 09-Feb-2026   (M/D/YY string)
        datetime(2026,10,2) → 10-Feb-2026  (Excel swapped 2/10/26)
    """
    if pd.isna(val):
        return pd.NaT

    # ── String input ─────────────────────────────────────────────────────
    if isinstance(val, str):
        val = val.strip()
        if not val:
            return pd.NaT
        parts = re.split(r'[/\-]', val)
        if len(parts) == 3:
            try:
                a, b, c = [int(p) for p in parts]
            except ValueError:
                return pd.to_datetime(val, dayfirst=True, errors="coerce")
            # Fix 2-digit year → 4-digit
            yr = c if c > 100 else (2000 + c)
            if a > 12:
                # First number > 12 → must be the day → DD/MM/YYYY
                return pd.Timestamp(year=yr, month=b, day=a)
            elif b > 12:
                # Second number > 12 → must be the day → MM/DD/YYYY
                return pd.Timestamp(year=yr, month=a, day=b)
            else:
                # Both ≤ 12 — ambiguous: use M/D/YY (US short, common in SF)
                try:
                    return pd.Timestamp(year=yr, month=a, day=b)
                except Exception:
                    return pd.Timestamp(year=yr, month=b, day=a)
        return pd.to_datetime(val, dayfirst=True, errors="coerce")

    # ── Datetime / Timestamp from Excel ──────────────────────────────────
    try:
        ts = pd.Timestamp(val)
        # Excel read '2/10/26' as 2026-10-02 (Oct 2) — should be Feb 10
        if ts.day <= 12 and ts.month > 2:
            return pd.Timestamp(year=ts.year, month=ts.day, day=ts.month)
        return ts
    except Exception:
        return pd.NaT


def fmt_date(val):
    """Format a date value to DD-MMM-YYYY. Returns '—' for NaT/None."""
    if pd.isna(val): return "—"
    try:
        ts = pd.Timestamp(val)
        return ts.strftime(DATE_FMT)
    except Exception:
        return str(val)


def clean_upload(df):
    df = df.loc[:, ~df.columns.str.startswith("Unnamed")].copy()
    rn = {}
    for c in df.columns:
        cl = c.strip()
        for s in SF_COLS + TEAM_COLS:
            if cl.lower().replace(" ","") == s.lower().replace(" ",""):
                rn[c] = s
    df.rename(columns=rn, inplace=True)
    if "Opportunity PAR" in df.columns:
        df["Opportunity PAR"] = df["Opportunity PAR"].apply(parse_par)
    if "Stage Duration" in df.columns:
        df["Stage Duration"] = pd.to_numeric(df["Stage Duration"], errors="coerce").fillna(0).astype(int)
    # Close Date uses US format (MM/DD/YYYY)
    if "Close Date" in df.columns:
        df["Close Date"] = pd.to_datetime(df["Close Date"], errors="coerce", dayfirst=False)
    # Received / Closed by Solutions use EU format (DD/MM/YYYY) — fix Excel swap
    if "Received by Solutions" in df.columns:
        df["Received by Solutions"] = df["Received by Solutions"].apply(fix_excel_eu_date)
    if "Closed by Solutions" in df.columns:
        df["Closed by Solutions"] = df["Closed by Solutions"].apply(fix_excel_eu_date)
    if "Status" in df.columns:  df["Status"] = df["Status"].fillna("Unassigned")
    if "Product" in df.columns: df["Product"] = df["Product"].fillna("General")
    return df.reset_index(drop=True)


def merge_masterfile(master, new_sf):
    for c in TEAM_COLS:
        if c not in master.columns: master[c] = ""
    old_o = set(master["Opportunity Name"].dropna())
    new_o = set(new_sf["Opportunity Name"].dropna())
    upd = 0
    for i, r in master.iterrows():
        o = r.get("Opportunity Name")
        if o in new_o:
            m = new_sf[new_sf["Opportunity Name"]==o].iloc[0]
            for c in SF_COLS:
                if c in new_sf.columns: master.at[i,c] = m[c]
            upd += 1
    added = new_o - old_o
    nr = new_sf[new_sf["Opportunity Name"].isin(added)].copy()
    for c in TEAM_COLS: nr[c] = ""
    master = pd.concat([master,nr], ignore_index=True)
    removed = old_o - new_o
    for o in removed:
        mask = master["Opportunity Name"]==o
        master.loc[mask,"Solutions Notes"] = master.loc[mask,"Solutions Notes"].fillna("").astype(str)+" [Removed from SF]"
    cols = [c for c in ALL_COLS if c in master.columns]
    return master[cols].reset_index(drop=True), {"updated":upd,"added":len(added),"removed":len(removed),"total":len(master)}


def fc(v):
    if pd.isna(v) or v==0: return "$0"
    if abs(v)>=1e6: return f"${v/1e6:,.1f}M"
    if abs(v)>=1e3: return f"${v/1e3:,.0f}K"
    return f"${v:,.0f}"


def pct(part, whole):
    return f"{part/whole*100:.0f}%" if whole else "0%"


def pl(fig, h=380, mb=40, mt=32):
    """Apply MBB-style layout to any Plotly figure — generous spacing."""
    fig.update_layout(
        font=dict(family="DM Sans, sans-serif", size=11.5, color=G800),
        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
        height=h, margin=dict(l=16, r=16, t=mt, b=mb),
        legend=dict(font=dict(size=9.5), orientation="h", y=-0.18, x=0.5, xanchor="center",
                    bgcolor="rgba(0,0,0,0)"),
    )
    fig.update_xaxes(gridcolor=G200, gridwidth=0.5, showline=True, linecolor=G200, linewidth=0.8,
                     tickfont=dict(size=10, color=G600))
    fig.update_yaxes(gridcolor=G200, gridwidth=0.5, showline=True, linecolor=G200, linewidth=0.8,
                     tickfont=dict(size=10, color=G600))
    return fig


def to_excel(df):
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        exp = df.copy()
        # Format date columns for export
        for dc in ["Close Date", "Received by Solutions", "Closed by Solutions"]:
            if dc in exp.columns:
                exp[dc] = pd.to_datetime(exp[dc], errors="coerce").apply(lambda x: x.strftime(DATE_FMT) if pd.notna(x) else "")
        exp.to_excel(w, index=False, sheet_name="Masterfile")
        ws = w.sheets["Masterfile"]
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        hf=PatternFill("solid",fgColor="002B49"); hw=Font(name="Calibri",bold=True,color="FFFFFF",size=10)
        tf=PatternFill("solid",fgColor="E8F5F3"); bd=Border(*(Side(style="thin",color="E1E4E8"),)*4)
        for ci,cn in enumerate(exp.columns,1):
            c=ws.cell(1,ci); c.fill,c.font,c.border=hf,hw,bd
            c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
            if cn in TEAM_COLS:
                for ri in range(2,len(exp)+2): ws.cell(ri,ci).fill=tf
            ml=max(len(str(cn)),*(len(str(x)) for x in exp[cn].head(50))) if len(exp) else len(str(cn))
            ws.column_dimensions[ws.cell(1,ci).column_letter].width=min(ml+4,42)
        bf=Font(name="Calibri",size=10)
        for ri in range(2,len(exp)+2):
            for ci in range(1,len(exp.columns)+1):
                c=ws.cell(ri,ci); c.font,c.border=bf,bd; c.alignment=Alignment(vertical="center",wrap_text=True)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════════
# SESSION STATE
# ═══════════════════════════════════════════════════════════════════════════════
if "master" not in st.session_state:
    st.session_state.master = None

# ═══════════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ═══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown(f"""
    <div style="padding:1.2rem 0 1rem; text-align:center;">
        <div style="font-size:.54rem; font-weight:700; letter-spacing:.26em; color:{GD}; text-transform:uppercase;">Marken · UPS Healthcare</div>
        <div style="font-size:.95rem; font-weight:700; color:{W}; margin-top:.2rem;">Solutions Pipeline</div>
        <div style="font-size:.62rem; color:rgba(255,255,255,0.30); margin-top:.12rem;">Precision Logistics</div>
    </div>
    """, unsafe_allow_html=True)
    st.divider()
    page = st.radio("", ["Dashboard","Masterfile Manager"], label_visibility="collapsed")
    st.divider()
    st.markdown(f'<div style="font-size:.56rem; color:rgba(255,255,255,0.22); text-align:center; line-height:1.5;">Report generated<br>{datetime.now().strftime("%d %b %Y · %H:%M")}</div>', unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
# HEADER
# ═══════════════════════════════════════════════════════════════════════════════
st.markdown(f"""
<div class="rh">
    <div><div class="rh-b">Marken · UPS Healthcare Precision Logistics</div>
        <div class="rh-t">Solutions Team — Global Pipeline Report</div>
        <div class="rh-s">Opportunity overview · Masterfile management · Analytics</div></div>
    <div class="rh-r"><div class="rh-d">{datetime.now().strftime(DATE_FMT)}</div><div class="rh-c">Confidential</div></div>
</div>
""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
# INITIAL UPLOAD
# ═══════════════════════════════════════════════════════════════════════════════
if st.session_state.master is None:
    st.markdown("""<div class="wf"><b>Getting started</b> — Upload your Salesforce export or existing Masterfile (.xlsx / .csv). The app detects the format and adds team columns if needed.</div>""", unsafe_allow_html=True)
    f = st.file_uploader("Upload Salesforce Export or Masterfile", type=["xlsx","xls","csv"], label_visibility="collapsed")
    if f:
        raw = pd.read_csv(f) if f.name.endswith(".csv") else pd.read_excel(f)
        cl = clean_upload(raw)
        for c in TEAM_COLS:
            if c not in cl.columns: cl[c] = ""
        cols = [c for c in ALL_COLS if c in cl.columns]
        st.session_state.master = cl[cols]
        st.rerun()
    st.stop()

df = st.session_state.master.copy()
df["Opportunity PAR"] = df["Opportunity PAR"].apply(parse_par)
df["Stage Duration"]  = pd.to_numeric(df.get("Stage Duration",0), errors="coerce").fillna(0).astype(int)
df["Close Date Parsed"] = pd.to_datetime(df["Close Date"], errors="coerce")

# Format display date columns
df["Close Date Display"] = df["Close Date Parsed"].apply(fmt_date)

if "Received by Solutions" in df.columns:
    df["Received by Solutions Parsed"] = pd.to_datetime(df["Received by Solutions"], errors="coerce")
    df["Received Display"] = df["Received by Solutions Parsed"].apply(fmt_date)
else:
    df["Received by Solutions Parsed"] = pd.NaT
    df["Received Display"] = "—"

if "Closed by Solutions" in df.columns:
    df["Closed by Solutions Parsed"] = pd.to_datetime(df["Closed by Solutions"], errors="coerce")
    df["Closed Display"] = df["Closed by Solutions Parsed"].apply(fmt_date)
else:
    df["Closed by Solutions Parsed"] = pd.NaT
    df["Closed Display"] = "—"

if "Status" not in df.columns: df["Status"] = "Unassigned"
else: df["Status"] = df["Status"].fillna("Unassigned")
if "Product" not in df.columns: df["Product"] = "General"
else: df["Product"] = df["Product"].fillna("General")

TODAY = pd.Timestamp.now().normalize()

# Solutions Cycle Time
if "Received by Solutions Parsed" in df.columns and "Closed by Solutions Parsed" in df.columns:
    df["Solutions Cycle Days"] = (df["Closed by Solutions Parsed"] - df["Received by Solutions Parsed"]).dt.days
else:
    df["Solutions Cycle Days"] = np.nan


# ═══════════════════════════════════════════════════════════════════════════════
#  DASHBOARD
# ═══════════════════════════════════════════════════════════════════════════════
if page == "Dashboard":

    # ── Filters ──────────────────────────────────────────────────────────────
    with st.expander("Filters", expanded=False):
        f1,f2,f3 = st.columns(3)
        sel_st = f1.multiselect("Stage", sorted(df["Stage"].dropna().unique()), default=sorted(df["Stage"].dropna().unique()))
        sel_sv = f2.multiselect("Service", sorted(df["Main Primary Service"].dropna().unique()), default=sorted(df["Main Primary Service"].dropna().unique()))
        sel_rg = f3.multiselect("Region", sorted(df["Owner Role"].dropna().unique()), default=sorted(df["Owner Role"].dropna().unique()))
        f4,f5,f6 = st.columns(3)
        sel_rs = f4.multiselect("Solution Resource", sorted(df["Solution Resource"].dropna().unique()), default=sorted(df["Solution Resource"].dropna().unique()))
        sel_status = f5.multiselect("Status", sorted(df["Status"].dropna().unique()), default=sorted(df["Status"].dropna().unique()))
        sel_product = f6.multiselect("Product", sorted(df["Product"].dropna().unique()), default=sorted(df["Product"].dropna().unique()))

    fdf = df[
        df["Stage"].isin(sel_st) & df["Main Primary Service"].isin(sel_sv) &
        df["Owner Role"].isin(sel_rg) & df["Solution Resource"].isin(sel_rs) &
        df["Status"].isin(sel_status) & df["Product"].isin(sel_product)
    ]

    total    = fdf["Opportunity PAR"].sum()
    n_opp    = len(fdf)
    n_cust   = fdf["Account Name"].nunique()
    n_svc    = fdf["Main Primary Service"].nunique()
    avg_deal = fdf["Opportunity PAR"].mean() if n_opp else 0
    avg_dur  = fdf["Stage Duration"].mean() if n_opp else 0
    past_due = fdf[fdf["Close Date Parsed"] < TODAY]
    aging_60 = fdf[fdf["Stage Duration"] > 60]

    n_working    = len(fdf[fdf["Status"]=="Working"])
    n_pending    = len(fdf[fdf["Status"]=="Pending"])
    n_unassigned = len(fdf[fdf["Status"]=="Unassigned"])
    n_products   = fdf[fdf["Product"]!="General"]["Product"].nunique()

    # Solutions date metrics (used in KPIs and Section 2)
    _rcv = fdf.dropna(subset=["Received by Solutions Parsed"]) if "Received by Solutions Parsed" in fdf.columns else pd.DataFrame()
    _cls = fdf.dropna(subset=["Closed by Solutions Parsed"]) if "Closed by Solutions Parsed" in fdf.columns else pd.DataFrame()
    n_received   = len(_rcv)
    pct_received = f"{n_received/n_opp*100:.0f}%" if n_opp else "0%"

    # ── KPIs ─────────────────────────────────────────────────────────────────
    st.markdown(f"""
    <div class="kr">
        <div class="kp" style="--ac:{TL};"><div class="kp-l">Pipeline Value</div><div class="kp-v">{fc(total)}</div><div class="kp-d">{n_opp} opportunities</div></div>
        <div class="kp" style="--ac:{NY};"><div class="kp-l">Customers</div><div class="kp-v">{n_cust}</div><div class="kp-d">Unique accounts</div></div>
        <div class="kp" style="--ac:{GD};"><div class="kp-l">Products in Scope</div><div class="kp-v">{n_svc}</div><div class="kp-d">{n_products} tagged product(s)</div></div>
        <div class="kp" style="--ac:{BA};"><div class="kp-l">Avg Deal Size</div><div class="kp-v">{fc(avg_deal)}</div><div class="kp-d">Per opportunity</div></div>
        <div class="kp" style="--ac:{TL};"><div class="kp-l">Solutions Received</div><div class="kp-v">{n_received}</div><div class="kp-d">{pct_received} of pipeline</div></div>
        <div class="kp" style="--ac:{G400};"><div class="kp-l">Avg Stage Duration</div><div class="kp-v">{avg_dur:.0f}d</div><div class="kp-d">{len(aging_60)} over 60 days</div></div>
        <div class="kp" style="--ac:{RD};"><div class="kp-l">Past Close Date</div><div class="kp-v">{len(past_due)}</div><div class="kp-d">{fc(past_due['Opportunity PAR'].sum())} at risk</div></div>
    </div>
    """, unsafe_allow_html=True)

    spacer("lg")

    # ══════════════════════════════════════════════════════════════════════════
    # 1 — PIPELINE OVERVIEW
    # ══════════════════════════════════════════════════════════════════════════
    st.markdown('<div class="sec">1 · Pipeline Overview</div>', unsafe_allow_html=True)
    spacer("md")

    p1a, _, p1b = st.columns([1.08, 0.08, 0.84])

    with p1a:
        st.markdown('<p class="so">Solutions Design holds the bulk of pipeline — most value remains in early stages</p>', unsafe_allow_html=True)
        sg = fdf.groupby("Stage").agg(Value=("Opportunity PAR","sum"), Count=("Opportunity Name","count")).reset_index()
        sg["StageOrd"] = sg["Stage"].apply(lambda x: STAGE_ORDER.index(x) if x in STAGE_ORDER else 99)
        sg = sg.sort_values("StageOrd")
        colors = []
        for s in sg["Stage"]:
            if "Closed/Lost" in s:    colors.append(RD)
            elif "Closed/Won" in s:   colors.append(GN)
            elif "Negotiation" in s:  colors.append(GD)
            elif "Proposal" in s:     colors.append(BA)
            elif "Design" in s:       colors.append(TL)
            else:                     colors.append(G400)
        fig = go.Figure(go.Bar(
            x=sg["Stage"], y=sg["Value"], marker_color=colors,
            text=[f"{fc(v)}<br><span style='font-size:9px;color:{G600}'>{c} opps</span>" for v,c in zip(sg["Value"],sg["Count"])],
            textposition="outside", textfont=dict(size=10.5),
        ))
        pl(fig, h=410, mb=16)
        fig.update_layout(showlegend=False, yaxis=dict(visible=False))
        fig.update_xaxes(showgrid=False, tickfont=dict(size=10))
        fig.update_yaxes(showgrid=False, showline=False)
        st.plotly_chart(fig, use_container_width=True)

    with p1b:
        st.markdown('<p class="so">Stage composition — value vs. count</p>', unsafe_allow_html=True)
        sg2 = sg.copy()
        fig2 = make_subplots(rows=1, cols=2, specs=[[{"type":"pie"},{"type":"pie"}]],
                             subplot_titles=["By Value","By Count"], horizontal_spacing=0.08)
        fig2.add_trace(go.Pie(labels=sg2["Stage"], values=sg2["Value"], hole=.52,
            marker=dict(colors=colors), textinfo="percent", textfont=dict(size=10),
            hovertemplate="<b>%{label}</b><br>$%{value:,.0f}<extra></extra>", sort=False), 1, 1)
        fig2.add_trace(go.Pie(labels=sg2["Stage"], values=sg2["Count"], hole=.52,
            marker=dict(colors=colors), textinfo="percent", textfont=dict(size=10),
            hovertemplate="<b>%{label}</b><br>%{value} opps<extra></extra>", sort=False), 1, 2)
        pl(fig2, h=410)
        fig2.update_layout(showlegend=False)
        fig2.update_annotations(font=dict(size=10, color=G600))
        st.plotly_chart(fig2, use_container_width=True)

    spacer("xl")

    # ══════════════════════════════════════════════════════════════════════════
    # 2 — STATUS & SOLUTIONS VELOCITY
    # ══════════════════════════════════════════════════════════════════════════
    st.markdown('<div class="sec">2 · Solutions Status & Velocity</div>', unsafe_allow_html=True)
    spacer("md")

    # ── Solutions Coverage KPIs ──────────────────────────────────────────
    rcv_col = _rcv
    cls_col = _cls
    n_closed   = len(cls_col)
    pct_closed = f"{n_closed/n_opp*100:.0f}%" if n_opp else "0%"
    cycle_days = fdf["Solutions Cycle Days"].dropna()
    avg_cycle  = f"{cycle_days.mean():.0f}d" if len(cycle_days) else "—"

    st.markdown(f"""
    <div class="kr">
        <div class="kp" style="--ac:{TL};"><div class="kp-l">Received by Solutions</div><div class="kp-v">{n_received}</div><div class="kp-d">{pct_received} of pipeline</div></div>
        <div class="kp" style="--ac:{GN};"><div class="kp-l">Closed by Solutions</div><div class="kp-v">{n_closed}</div><div class="kp-d">{pct_closed} of pipeline</div></div>
        <div class="kp" style="--ac:{BA};"><div class="kp-l">Avg Cycle Time</div><div class="kp-v">{avg_cycle}</div><div class="kp-d">Received → Closed</div></div>
        <div class="kp" style="--ac:{GN};"><div class="kp-l">Status: Working</div><div class="kp-v">{n_working}</div><div class="kp-d">{n_pending} pending · {n_unassigned} unassigned</div></div>
    </div>
    """, unsafe_allow_html=True)

    spacer("md")

    # ── Row 1: Status bar + Status donut ─────────────────────────────────
    s2a, _, s2b = st.columns([0.52, 0.06, 0.42])

    with s2a:
        st.markdown('<p class="so">Opportunity status distribution — workload snapshot</p>', unsafe_allow_html=True)
        stat_g = fdf.groupby("Status").agg(Value=("Opportunity PAR","sum"), Count=("Opportunity Name","count")).reset_index().sort_values("Count", ascending=False)
        stat_colors = [STATUS_COLORS.get(s, G400) for s in stat_g["Status"]]
        fig_stat = go.Figure(go.Bar(
            x=stat_g["Status"], y=stat_g["Count"], marker_color=stat_colors,
            text=[f"{c}<br><span style='font-size:9px;color:{G600}'>{fc(v)}</span>" for c,v in zip(stat_g["Count"],stat_g["Value"])],
            textposition="outside", textfont=dict(size=11),
        ))
        pl(fig_stat, h=380, mb=16)
        fig_stat.update_layout(showlegend=False, yaxis=dict(visible=False), bargap=0.4)
        fig_stat.update_xaxes(showgrid=False)
        fig_stat.update_yaxes(showgrid=False, showline=False)
        st.plotly_chart(fig_stat, use_container_width=True)

    with s2b:
        st.markdown('<p class="so">Status × pipeline value share</p>', unsafe_allow_html=True)
        fig_stat2 = go.Figure(go.Pie(
            labels=stat_g["Status"], values=stat_g["Value"], hole=.55,
            marker=dict(colors=stat_colors),
            textinfo="label+percent", textfont=dict(size=10.5),
            hovertemplate="<b>%{label}</b><br>$%{value:,.0f}<br>%{percent}<extra></extra>",
        ))
        pl(fig_stat2, h=380)
        fig_stat2.update_layout(showlegend=False)
        st.plotly_chart(fig_stat2, use_container_width=True)

    spacer("lg")

    # ── Row 2: Solutions intake timeline + Lifecycle view ────────────────
    s2c, _, s2d = st.columns([0.52, 0.06, 0.42])

    with s2c:
        st.markdown('<p class="so">Solutions intake timeline — when opportunities were received</p>', unsafe_allow_html=True)
        if "Received by Solutions Parsed" in fdf.columns and len(rcv_col):
            rcv_t = rcv_col.copy()
            rcv_t["Rcv Week"] = rcv_t["Received by Solutions Parsed"].dt.to_period("W").dt.to_timestamp()
            wk = rcv_t.groupby("Rcv Week").agg(Count=("Opportunity Name","count"), Value=("Opportunity PAR","sum")).reset_index().sort_values("Rcv Week")
            fig_rcv = go.Figure()
            fig_rcv.add_trace(go.Bar(x=wk["Rcv Week"], y=wk["Count"], marker_color=TL, name="Received",
                text=[f"{c}<br><span style='font-size:9px;color:{G600}'>{fc(v)}</span>" for c,v in zip(wk["Count"],wk["Value"])],
                textposition="outside", textfont=dict(size=10.5)))
            pl(fig_rcv, h=380, mb=20)
            fig_rcv.update_layout(showlegend=False, yaxis=dict(visible=False), bargap=0.35,
                xaxis=dict(tickformat="%d-%b-%Y", tickfont=dict(size=9.5)))
            fig_rcv.update_xaxes(showgrid=False)
            fig_rcv.update_yaxes(showgrid=False, showline=False)
            st.plotly_chart(fig_rcv, use_container_width=True)
        else:
            st.info("No 'Received by Solutions' dates populated yet.")

    with s2d:
        st.markdown('<p class="so">Solutions lifecycle — received vs. closed per opportunity</p>', unsafe_allow_html=True)
        if "Received by Solutions Parsed" in fdf.columns and len(rcv_col):
            lf = rcv_col[["Account Name","Received by Solutions Parsed","Closed by Solutions Parsed",
                          "Opportunity PAR","Status"]].copy()
            lf = lf.sort_values("Received by Solutions Parsed")
            lf["End"] = lf["Closed by Solutions Parsed"].fillna(TODAY)
            lf["Acct Short"] = lf["Account Name"].apply(lambda x: (x[:22]+"…") if len(str(x))>22 else x)
            lf["Is Closed"] = lf["Closed by Solutions Parsed"].notna()

            fig_lf = go.Figure()
            for idx, row in lf.iterrows():
                clr = GN if row["Is Closed"] else TL
                fig_lf.add_trace(go.Bar(
                    y=[row["Acct Short"]], x=[(row["End"]-row["Received by Solutions Parsed"]).days],
                    base=[row["Received by Solutions Parsed"]],
                    orientation="h", marker_color=clr, marker_line=dict(width=0),
                    showlegend=False,
                    hovertemplate=(
                        f"<b>{row['Account Name']}</b><br>"
                        f"Received: {fmt_date(row['Received by Solutions Parsed'])}<br>"
                        f"{'Closed: '+fmt_date(row['Closed by Solutions Parsed']) if row['Is Closed'] else 'Open (in progress)'}<br>"
                        f"Value: {fc(row['Opportunity PAR'])}<extra></extra>"
                    ),
                ))
            # Add legend manually
            fig_lf.add_trace(go.Bar(y=[None], x=[None], marker_color=GN, name="Closed", showlegend=True))
            fig_lf.add_trace(go.Bar(y=[None], x=[None], marker_color=TL, name="Open", showlegend=True))
            pl(fig_lf, h=max(380, 48*len(lf)), mb=20)
            fig_lf.update_layout(
                barmode="overlay", yaxis=dict(tickfont=dict(size=9.5), autorange="reversed"),
                xaxis=dict(tickformat="%d-%b-%Y", tickfont=dict(size=9)),
                legend=dict(font=dict(size=9), orientation="h", y=-0.12, x=0.5, xanchor="center"),
            )
            fig_lf.update_xaxes(showgrid=True, gridcolor=G200, gridwidth=0.4, showline=False)
            fig_lf.update_yaxes(showgrid=False, showline=False)
            st.plotly_chart(fig_lf, use_container_width=True)
        else:
            st.info("No lifecycle data available yet.")

    spacer("lg")

    # ── Resource × Status heatmap ────────────────────────────────────────
    st.markdown('<p class="so">Solution Resource workload by status — identify capacity constraints and bottlenecks</p>', unsafe_allow_html=True)
    rs_stat = pd.crosstab(fdf["Solution Resource"], fdf["Status"], values=fdf["Opportunity Name"], aggfunc="count").fillna(0).astype(int)
    status_order = ["Working","Pending","Unassigned"]
    rs_stat = rs_stat[[c for c in status_order if c in rs_stat.columns] + [c for c in rs_stat.columns if c not in status_order]]
    fig_rs = go.Figure(go.Heatmap(
        z=rs_stat.values, x=rs_stat.columns.tolist(), y=rs_stat.index.tolist(),
        colorscale=[[0,W],[.25,TLL],[.6,TL],[1,NY]],
        text=rs_stat.values, texttemplate="%{text}", textfont=dict(size=12.5),
        hovertemplate="Resource: %{y}<br>Status: %{x}<br>Count: %{z}<extra></extra>",
        showscale=False, xgap=4, ygap=4,
    ))
    pl(fig_rs, h=max(260, 48*len(rs_stat)), mt=8)
    fig_rs.update_layout(xaxis=dict(tickfont=dict(size=10.5), side="top"), yaxis=dict(tickfont=dict(size=10.5), autorange="reversed"))
    fig_rs.update_xaxes(showgrid=False, showline=False)
    fig_rs.update_yaxes(showgrid=False, showline=False)
    st.plotly_chart(fig_rs, use_container_width=True)

    spacer("xl")

    # ══════════════════════════════════════════════════════════════════════════
    # 3 — PRODUCT ANALYSIS
    # ══════════════════════════════════════════════════════════════════════════
    st.markdown('<div class="sec">3 · Product Segmentation</div>', unsafe_allow_html=True)
    spacer("md")

    p3a, _, p3b = st.columns([0.48, 0.06, 0.46])

    with p3a:
        prod_g = fdf.groupby("Product").agg(Value=("Opportunity PAR","sum"), Count=("Opportunity Name","count"),
            Avg=("Opportunity PAR","mean")).reset_index().sort_values("Value", ascending=False)
        top_prod = prod_g.iloc[0]["Product"] if len(prod_g) else "N/A"
        top_prod_pct = prod_g.iloc[0]["Value"]/prod_g["Value"].sum()*100 if len(prod_g) and prod_g["Value"].sum()>0 else 0
        st.markdown(f'<p class="so">{top_prod} represents {top_prod_pct:.0f}% of pipeline by value</p>', unsafe_allow_html=True)
        prod_colors = [SEQ[i%len(SEQ)] for i in range(len(prod_g))]
        fig_prod = go.Figure(go.Bar(
            x=prod_g["Product"], y=prod_g["Value"], marker_color=prod_colors,
            text=[f"{fc(v)}<br><span style='font-size:9px;color:{G600}'>{c} opps</span>" for v,c in zip(prod_g["Value"],prod_g["Count"])],
            textposition="outside", textfont=dict(size=10.5),
        ))
        pl(fig_prod, h=400, mb=16)
        fig_prod.update_layout(showlegend=False, yaxis=dict(visible=False), bargap=0.35)
        fig_prod.update_xaxes(showgrid=False, tickfont=dict(size=10.5))
        fig_prod.update_yaxes(showgrid=False, showline=False)
        st.plotly_chart(fig_prod, use_container_width=True)

    with p3b:
        st.markdown('<p class="so">Product × Region — strategic coverage</p>', unsafe_allow_html=True)
        pr_ht = pd.crosstab(fdf["Product"], fdf["Owner Role"], values=fdf["Opportunity PAR"], aggfunc="sum").fillna(0)
        fig_pr = go.Figure(go.Heatmap(
            z=pr_ht.values, x=pr_ht.columns.tolist(), y=pr_ht.index.tolist(),
            colorscale=[[0,W],[.25,TLL],[.6,TL],[1,NY]],
            text=[[fc(v) for v in row] for row in pr_ht.values], texttemplate="%{text}", textfont=dict(size=11.5),
            hovertemplate="Product: %{y}<br>Region: %{x}<br>Value: $%{z:,.0f}<extra></extra>",
            showscale=False, xgap=4, ygap=4,
        ))
        pl(fig_pr, h=max(280, 66*len(pr_ht)), mt=8)
        fig_pr.update_layout(xaxis=dict(tickfont=dict(size=10.5), side="top"), yaxis=dict(tickfont=dict(size=10.5), autorange="reversed"))
        fig_pr.update_xaxes(showgrid=False, showline=False)
        fig_pr.update_yaxes(showgrid=False, showline=False)
        st.plotly_chart(fig_pr, use_container_width=True)

    spacer("lg")

    # Product × Service
    st.markdown('<p class="so">Product × Service cross-reference — identifying service-product alignment</p>', unsafe_allow_html=True)
    ps_ht = pd.crosstab(fdf["Product"], fdf["Main Primary Service"], values=fdf["Opportunity PAR"], aggfunc="sum").fillna(0)
    fig_ps = go.Figure(go.Heatmap(
        z=ps_ht.values, x=ps_ht.columns.tolist(), y=ps_ht.index.tolist(),
        colorscale=[[0,W],[.25,"#FFF8E1"],[.6,GD],[1,NY]],
        text=[[fc(v) for v in row] for row in ps_ht.values], texttemplate="%{text}", textfont=dict(size=10.5),
        hovertemplate="Product: %{y}<br>Service: %{x}<br>Value: $%{z:,.0f}<extra></extra>",
        showscale=False, xgap=4, ygap=4,
    ))
    pl(fig_ps, h=max(260, 66*len(ps_ht)), mt=8)
    fig_ps.update_layout(xaxis=dict(tickfont=dict(size=10), side="top"), yaxis=dict(tickfont=dict(size=10.5), autorange="reversed"))
    fig_ps.update_xaxes(showgrid=False, showline=False)
    fig_ps.update_yaxes(showgrid=False, showline=False)
    st.plotly_chart(fig_ps, use_container_width=True)

    spacer("xl")

    # ══════════════════════════════════════════════════════════════════════════
    # 4 — CUSTOMER ANALYSIS
    # ══════════════════════════════════════════════════════════════════════════
    st.markdown('<div class="sec">4 · Customer Concentration</div>', unsafe_allow_html=True)
    spacer("md")

    c2a, _, c2b = st.columns([1.08, 0.08, 0.84])

    with c2a:
        cu = fdf.groupby("Account Name").agg(Value=("Opportunity PAR","sum"), Count=("Opportunity Name","count")).reset_index().sort_values("Value", ascending=False)
        top5_pct = cu.head(5)["Value"].sum()/cu["Value"].sum()*100 if cu["Value"].sum()>0 else 0
        st.markdown(f'<p class="so">Top 5 accounts represent {top5_pct:.0f}% of total pipeline — concentration risk to monitor</p>', unsafe_allow_html=True)
        cu_top = cu.head(10).sort_values("Value", ascending=True)
        fig3 = go.Figure(go.Bar(
            y=cu_top["Account Name"], x=cu_top["Value"], orientation="h",
            marker=dict(color=cu_top["Value"], colorscale=[[0,"#B2DFDB"],[.5,TL],[1,NY]], showscale=False),
            text=[f"  {fc(v)}  ({c})" for v,c in zip(cu_top["Value"],cu_top["Count"])],
            textposition="outside", textfont=dict(size=10.5, color=G800),
        ))
        pl(fig3, h=max(380, 44*len(cu_top)))
        fig3.update_layout(showlegend=False, xaxis=dict(visible=False), yaxis=dict(tickfont=dict(size=10)))
        fig3.update_xaxes(showgrid=False, showline=False)
        fig3.update_yaxes(showgrid=False, showline=False)
        st.plotly_chart(fig3, use_container_width=True)

    with c2b:
        st.markdown('<p class="so">Cumulative concentration (Pareto curve)</p>', unsafe_allow_html=True)
        cu_sorted = cu.sort_values("Value", ascending=False).reset_index(drop=True)
        cu_sorted["CumVal"] = cu_sorted["Value"].cumsum()
        cu_sorted["CumPct"] = cu_sorted["CumVal"]/cu_sorted["Value"].sum()*100
        cu_sorted["Rank"]   = range(1, len(cu_sorted)+1)
        fig4 = go.Figure()
        fig4.add_trace(go.Bar(x=cu_sorted["Rank"], y=cu_sorted["Value"], marker_color=TL, name="Individual Value",
            hovertemplate="<b>%{customdata}</b><br>$%{y:,.0f}<extra></extra>", customdata=cu_sorted["Account Name"]))
        fig4.add_trace(go.Scatter(x=cu_sorted["Rank"], y=cu_sorted["CumPct"], mode="lines+markers",
            line=dict(color=NY, width=2.5), marker=dict(size=5, color=NY), name="Cumulative %", yaxis="y2"))
        fig4.add_hline(y=80, line_dash="dot", line_color=RD, opacity=0.45, annotation_text="80 %",
            annotation_position="right", annotation_font=dict(size=9, color=RD), yref="y2")
        pl(fig4, h=max(380, 44*len(cu_top)), mb=50)
        fig4.update_layout(
            yaxis=dict(visible=False),
            yaxis2=dict(title="Cumulative %", titlefont=dict(size=9), side="right", overlaying="y", range=[0,108], showgrid=False),
            xaxis=dict(title="Customer Rank", titlefont=dict(size=9.5), tickfont=dict(size=9.5)),
            legend=dict(font=dict(size=8.5)),
        )
        fig4.update_xaxes(showgrid=False)
        fig4.update_yaxes(showgrid=False, showline=False)
        st.plotly_chart(fig4, use_container_width=True)

    spacer("xl")

    # ══════════════════════════════════════════════════════════════════════════
    # 5 — PRODUCT / SERVICE MIX
    # ══════════════════════════════════════════════════════════════════════════
    st.markdown('<div class="sec">5 · Product / Service Mix</div>', unsafe_allow_html=True)
    spacer("md")

    s3a, _, s3b = st.columns([0.50, 0.06, 0.44])

    with s3a:
        sv = fdf.groupby("Main Primary Service").agg(Value=("Opportunity PAR","sum"), Count=("Opportunity Name","count"),
            Avg=("Opportunity PAR","mean")).reset_index().sort_values("Value", ascending=False)
        top_svc = sv.iloc[0]["Main Primary Service"] if len(sv) else "N/A"
        top_svc_pct = sv.iloc[0]["Value"]/sv["Value"].sum()*100 if len(sv) and sv["Value"].sum()>0 else 0
        st.markdown(f'<p class="so">{top_svc} accounts for {top_svc_pct:.0f}% of pipeline value by service</p>', unsafe_allow_html=True)
        fig5 = go.Figure(go.Pie(
            labels=sv["Main Primary Service"], values=sv["Value"], hole=.52,
            marker=dict(colors=SEQ[:len(sv)]),
            textinfo="label+percent", textfont=dict(size=10.5),
            hovertemplate="<b>%{label}</b><br>$%{value:,.0f}<br>%{percent}<extra></extra>", sort=True,
        ))
        pl(fig5, h=400)
        fig5.update_layout(showlegend=False)
        st.plotly_chart(fig5, use_container_width=True)

    with s3b:
        st.markdown('<p class="so">Average deal size by service type</p>', unsafe_allow_html=True)
        sv_s = sv.sort_values("Avg", ascending=True)
        fig6 = go.Figure(go.Bar(
            y=sv_s["Main Primary Service"], x=sv_s["Avg"], orientation="h", marker_color=NY,
            text=[f"  {fc(v)}  ({c} opps)" for v,c in zip(sv_s["Avg"],sv_s["Count"])],
            textposition="outside", textfont=dict(size=10, color=G800),
        ))
        pl(fig6, h=400)
        fig6.update_layout(showlegend=False, xaxis=dict(visible=False), yaxis=dict(tickfont=dict(size=10)))
        fig6.update_xaxes(showgrid=False, showline=False)
        fig6.update_yaxes(showgrid=False, showline=False)
        st.plotly_chart(fig6, use_container_width=True)

    spacer("lg")

    # Service × Region Heatmap
    st.markdown('<p class="so">Service demand mapped by region</p>', unsafe_allow_html=True)
    ht = pd.crosstab(fdf["Main Primary Service"], fdf["Owner Role"], values=fdf["Opportunity PAR"], aggfunc="sum").fillna(0)
    fig7 = go.Figure(go.Heatmap(
        z=ht.values, x=ht.columns.tolist(), y=ht.index.tolist(),
        colorscale=[[0,W],[.25,TLL],[.6,TL],[1,NY]],
        text=[[fc(v) for v in row] for row in ht.values], texttemplate="%{text}", textfont=dict(size=10.5),
        hovertemplate="Service: %{y}<br>Region: %{x}<br>Value: $%{z:,.0f}<extra></extra>",
        showscale=False, xgap=4, ygap=4,
    ))
    pl(fig7, h=max(280, 48*len(ht)), mt=8)
    fig7.update_layout(xaxis=dict(tickfont=dict(size=10.5), side="top"), yaxis=dict(tickfont=dict(size=10.5), autorange="reversed"))
    fig7.update_xaxes(showgrid=False, showline=False)
    fig7.update_yaxes(showgrid=False, showline=False)
    st.plotly_chart(fig7, use_container_width=True)

    spacer("xl")

    # ══════════════════════════════════════════════════════════════════════════
    # 6 — REGIONAL ANALYSIS
    # ══════════════════════════════════════════════════════════════════════════
    st.markdown('<div class="sec">6 · Regional Split</div>', unsafe_allow_html=True)
    spacer("md")

    r4a, _, r4b = st.columns([1, 0.06, 1])

    with r4a:
        rg = fdf.groupby("Owner Role").agg(Value=("Opportunity PAR","sum"), Count=("Opportunity Name","count"),
            Avg=("Opportunity PAR","mean"), AvgDur=("Stage Duration","mean")).reset_index().sort_values("Value", ascending=False)
        top_reg = rg.iloc[0]["Owner Role"] if len(rg) else "N/A"
        st.markdown(f'<p class="so">{top_reg} leads the pipeline by value</p>', unsafe_allow_html=True)
        fig8 = go.Figure(go.Bar(
            x=rg["Owner Role"], y=rg["Value"],
            marker_color=[NY,TL,GD,BA,G600][:len(rg)],
            text=[f"{fc(v)}<br><span style='font-size:9px'>{c} opps · Avg {fc(a)}</span>" for v,c,a in zip(rg["Value"],rg["Count"],rg["Avg"])],
            textposition="inside", textfont=dict(size=11, color=W),
        ))
        pl(fig8, h=390)
        fig8.update_layout(showlegend=False, yaxis=dict(visible=False), bargap=0.35)
        fig8.update_xaxes(showgrid=False, tickfont=dict(size=10.5))
        fig8.update_yaxes(showgrid=False, showline=False)
        st.plotly_chart(fig8, use_container_width=True)

    with r4b:
        st.markdown('<p class="so">Region × stage: where is each region in the pipeline?</p>', unsafe_allow_html=True)
        rs_ht = pd.crosstab(fdf["Owner Role"], fdf["Stage"], values=fdf["Opportunity PAR"], aggfunc="sum").fillna(0)
        ordered = [s for s in STAGE_ORDER if s in rs_ht.columns]
        extra   = [s for s in rs_ht.columns if s not in STAGE_ORDER]
        rs_ht   = rs_ht[ordered+extra]
        stg_colors = {"Information Gathering":G400, "Solutions Design":TL, "Proposal/Price Quote":BA,
                      "Proposal Price/Quote":BA, "Negotiations":GD, "Closed/Won":GN, "Closed/Lost":RD}
        fig9 = go.Figure()
        for col in rs_ht.columns:
            fig9.add_trace(go.Bar(
                x=rs_ht.index, y=rs_ht[col], name=col, marker_color=stg_colors.get(col, G400),
                text=[fc(v) if v>0 else "" for v in rs_ht[col]], textposition="inside", textfont=dict(size=9, color=W),
            ))
        pl(fig9, h=390, mb=16)
        fig9.update_layout(barmode="stack", yaxis=dict(visible=False), bargap=0.3, legend=dict(font=dict(size=8.5)))
        fig9.update_xaxes(showgrid=False, tickfont=dict(size=10.5))
        fig9.update_yaxes(showgrid=False, showline=False)
        st.plotly_chart(fig9, use_container_width=True)

    spacer("xl")

    # ══════════════════════════════════════════════════════════════════════════
    # 7 — SOLUTION RESOURCE WORKLOAD
    # ══════════════════════════════════════════════════════════════════════════
    st.markdown('<div class="sec">7 · Solution Resource Workload</div>', unsafe_allow_html=True)
    spacer("md")

    rw = fdf.groupby("Solution Resource").agg(Value=("Opportunity PAR","sum"), Count=("Opportunity Name","count"),
        AvgDur=("Stage Duration","mean"), Cust=("Account Name","nunique")).reset_index().sort_values("Value", ascending=False)

    r5a, _, r5b = st.columns([0.56, 0.06, 0.38])

    with r5a:
        st.markdown(f'<p class="so">{rw.iloc[0]["Solution Resource"]} carries the largest pipeline at {fc(rw.iloc[0]["Value"])}</p>' if len(rw) else '', unsafe_allow_html=True)
        fig10 = go.Figure()
        fig10.add_trace(go.Bar(x=rw["Solution Resource"], y=rw["Count"], name="# Opportunities", marker_color=NY,
            text=rw["Count"], textposition="auto", textfont=dict(color=W, size=10.5)))
        fig10.add_trace(go.Scatter(x=rw["Solution Resource"], y=rw["Value"], name="Total Value ($)",
            mode="markers+lines", marker=dict(color=GD, size=10, line=dict(width=1.5, color=NY)),
            line=dict(color=GD, width=2.5), yaxis="y2"))
        pl(fig10, h=400, mb=55)
        fig10.update_layout(
            yaxis=dict(title="# Opps", titlefont=dict(size=9.5), side="left"),
            yaxis2=dict(title="Value ($)", titlefont=dict(size=9.5), side="right", overlaying="y", showgrid=False),
            xaxis=dict(tickangle=15, tickfont=dict(size=9.5)), bargap=0.3,
        )
        st.plotly_chart(fig10, use_container_width=True)

    with r5b:
        st.markdown('<p class="so">Resource detail breakdown</p>', unsafe_allow_html=True)
        rw_status = fdf.groupby("Solution Resource")["Status"].value_counts().unstack(fill_value=0)
        rw_merged = rw.set_index("Solution Resource").join(rw_status, how="left").reset_index()
        rd = rw_merged.rename(columns={"Solution Resource":"Resource","Count":"Opps","Value":"Pipeline","AvgDur":"Avg Days","Cust":"Accounts"}).copy()
        rd["Pipeline"] = rd["Pipeline"].apply(lambda x: f"${x:,.0f}")
        rd["Avg Days"] = rd["Avg Days"].apply(lambda x: f"{x:.0f}")
        display_cols = ["Resource","Opps","Pipeline","Avg Days","Accounts"]
        for sc in ["Working","Pending","Unassigned"]:
            if sc in rd.columns:
                rd[sc] = rd[sc].astype(int)
                display_cols.append(sc)
        st.dataframe(rd[display_cols], use_container_width=True, height=400, hide_index=True)

    spacer("lg")

    # Resource × Region
    st.markdown('<p class="so">Resource allocation by region</p>', unsafe_allow_html=True)
    rr = pd.crosstab(fdf["Solution Resource"], fdf["Owner Role"], values=fdf["Opportunity Name"], aggfunc="count").fillna(0).astype(int)
    fig11 = go.Figure(go.Heatmap(
        z=rr.values, x=rr.columns.tolist(), y=rr.index.tolist(),
        colorscale=[[0,W],[.4,TLL],[1,NY]],
        text=rr.values, texttemplate="%{text}", textfont=dict(size=12),
        showscale=False, xgap=4, ygap=4,
    ))
    pl(fig11, h=max(260, 46*len(rr)), mt=8)
    fig11.update_layout(xaxis=dict(tickfont=dict(size=10.5), side="top"), yaxis=dict(tickfont=dict(size=10.5), autorange="reversed"))
    fig11.update_xaxes(showgrid=False, showline=False)
    fig11.update_yaxes(showgrid=False, showline=False)
    st.plotly_chart(fig11, use_container_width=True)

    spacer("xl")

    # ══════════════════════════════════════════════════════════════════════════
    # 8 — TIMELINE & AGING
    # ══════════════════════════════════════════════════════════════════════════
    st.markdown('<div class="sec">8 · Timeline & Aging Analysis</div>', unsafe_allow_html=True)
    spacer("md")

    t6a, _, t6b = st.columns([1, 0.06, 1])

    with t6a:
        st.markdown('<p class="so">Expected revenue by close month</p>', unsafe_allow_html=True)
        tl = fdf.dropna(subset=["Close Date Parsed"]).copy()
        tl["Month"] = tl["Close Date Parsed"].dt.to_period("M").dt.to_timestamp()
        mo = tl.groupby("Month").agg(Value=("Opportunity PAR","sum"), Count=("Opportunity Name","count")).reset_index().sort_values("Month")
        fig12 = go.Figure()
        fig12.add_trace(go.Bar(x=mo["Month"], y=mo["Value"], marker_color=TL, name="Pipeline Value",
            text=[f"{fc(v)}" for v in mo["Value"]], textposition="outside", textfont=dict(size=9.5)))
        fig12.add_trace(go.Scatter(x=mo["Month"], y=mo["Count"], mode="markers+lines",
            marker=dict(color=NY, size=7), line=dict(color=NY, width=2), name="# Opps", yaxis="y2"))
        pl(fig12, h=400, mb=50)
        fig12.update_layout(
            yaxis=dict(visible=False),
            yaxis2=dict(title="# Opps", titlefont=dict(size=9.5), side="right", overlaying="y", showgrid=False),
            xaxis=dict(tickformat="%b '%y", tickfont=dict(size=9.5)),
        )
        fig12.update_xaxes(showgrid=False)
        fig12.update_yaxes(showgrid=False, showline=False)
        st.plotly_chart(fig12, use_container_width=True)

    with t6b:
        st.markdown('<p class="so">Stage duration distribution — identify outliers and bottlenecks</p>', unsafe_allow_html=True)
        stages_for_box = [s for s in STAGE_ORDER if s in fdf["Stage"].unique()]
        fig13 = go.Figure()
        for s in stages_for_box:
            sd = fdf[fdf["Stage"]==s]
            fig13.add_trace(go.Box(y=sd["Stage Duration"], name=s, marker_color=TL, boxmean=True,
                fillcolor=TLL, line=dict(color=TL)))
        pl(fig13, h=400)
        fig13.update_layout(showlegend=False, yaxis=dict(title="Days", titlefont=dict(size=9.5)),
                            xaxis=dict(tickfont=dict(size=9.5)))
        st.plotly_chart(fig13, use_container_width=True)

    spacer("lg")

    # Bubble chart
    st.markdown('<p class="so">Opportunity landscape: value vs. stage duration — bubble size = deal value, color = stage</p>', unsafe_allow_html=True)
    bdf = fdf.dropna(subset=["Close Date Parsed"]).copy()
    stg_c = {"Information Gathering":G400, "Solutions Design":TL, "Proposal/Price Quote":BA,
             "Proposal Price/Quote":BA, "Negotiations":GD, "Closed/Won":GN, "Closed/Lost":RD}
    fig14 = go.Figure()
    for stg in bdf["Stage"].unique():
        sd = bdf[bdf["Stage"]==stg]
        fig14.add_trace(go.Scatter(
            x=sd["Stage Duration"], y=sd["Opportunity PAR"], mode="markers", name=stg,
            marker=dict(size=sd["Opportunity PAR"].apply(lambda x: max(8, min(40, x/70000))),
                        color=stg_c.get(stg, G400), line=dict(width=1.2, color=W), opacity=0.82),
            text=sd["Account Name"],
            hovertemplate="<b>%{text}</b><br>Duration: %{x}d<br>Value: $%{y:,.0f}<extra></extra>",
        ))
    fig14.add_vline(x=60, line_dash="dot", line_color=RD, opacity=0.35,
                    annotation_text="60-day threshold", annotation_position="top",
                    annotation_font=dict(size=9, color=RD))
    pl(fig14, h=440, mb=50)
    fig14.update_layout(
        xaxis=dict(title="Stage Duration (days)", titlefont=dict(size=10.5)),
        yaxis=dict(title="PAR Value ($)", titlefont=dict(size=10.5)),
        legend=dict(font=dict(size=9)),
    )
    st.plotly_chart(fig14, use_container_width=True)

    spacer("lg")

    # Aging table
    st.markdown('<p class="so">Aging flags — opportunities requiring attention</p>', unsafe_allow_html=True)
    ag = fdf.copy()
    ag["Flag"] = ""
    ag.loc[ag["Stage Duration"]>90, "Flag"] = "🔴 >90d"
    ag.loc[(ag["Stage Duration"]>60)&(ag["Stage Duration"]<=90), "Flag"] = "🟡 >60d"
    ag.loc[ag["Close Date Parsed"]<TODAY, "Flag"] = ag.loc[ag["Close Date Parsed"]<TODAY, "Flag"].astype(str)+" ⚠ Past Due"
    ag_flagged = ag[ag["Flag"].str.len()>0].sort_values("Stage Duration", ascending=False)
    if len(ag_flagged):
        disp_ag = ag_flagged[["Flag","Account Name","Opportunity Name","Stage","Status","Product",
                              "Opportunity PAR","Stage Duration","Close Date Display","Notes"]].copy()
        disp_ag = disp_ag.rename(columns={"Close Date Display": "Close Date"})
        st.dataframe(
            disp_ag.style.format({"Opportunity PAR":"${:,.0f}"}),
            use_container_width=True, height=min(380, 38*len(ag_flagged)+38), hide_index=True,
        )
    else:
        st.success("No aging flags — all opportunities within normal parameters.")

    spacer("xl")

    # ══════════════════════════════════════════════════════════════════════════
    # 9 — RISK & CLOSED/LOST
    # ══════════════════════════════════════════════════════════════════════════
    st.markdown('<div class="sec">9 · Risk & Attention Items</div>', unsafe_allow_html=True)
    spacer("md")

    r7a, _, r7b = st.columns([1, 0.06, 1])

    with r7a:
        lost = fdf[fdf["Stage"]=="Closed/Lost"]
        if len(lost):
            st.markdown(f'<div class="al al-r"><b>{len(lost)} Closed/Lost</b> opportunities totaling <b>{fc(lost["Opportunity PAR"].sum())}</b></div>', unsafe_allow_html=True)
            spacer("md")
            for _, r in lost.iterrows():
                st.markdown(f"""
                <div style="background:{W}; border:1px solid {G200}; border-radius:5px; padding:.8rem 1rem; margin-bottom:.5rem; font-size:.78rem;">
                    <b style="color:{NY};">{r['Account Name']}</b> · {fc(r['Opportunity PAR'])}<br>
                    <span style="color:{G600};">{r['Notes'] if pd.notna(r['Notes']) else 'No notes'}</span>
                </div>
                """, unsafe_allow_html=True)
        else:
            st.success("No Closed/Lost opportunities in current view.")

    with r7b:
        st.markdown('<p class="so">Largest deals at risk (high value × high duration)</p>', unsafe_allow_html=True)
        risk = fdf[~fdf["Stage"].str.contains("Closed", na=False)].copy()
        risk["RiskScore"] = risk["Opportunity PAR"] * np.log1p(risk["Stage Duration"])
        risk_top = risk.nlargest(5, "RiskScore")
        if len(risk_top):
            fig15 = go.Figure(go.Bar(
                y=risk_top["Account Name"], x=risk_top["Opportunity PAR"], orientation="h",
                marker_color=[RD if d>60 else GD if d>30 else TL for d in risk_top["Stage Duration"]],
                text=[f"  {fc(v)} · {d}d" for v,d in zip(risk_top["Opportunity PAR"],risk_top["Stage Duration"])],
                textposition="outside", textfont=dict(size=10.5, color=G800),
            ))
            pl(fig15, h=300)
            fig15.update_layout(showlegend=False, xaxis=dict(visible=False), yaxis=dict(tickfont=dict(size=10)))
            fig15.update_xaxes(showgrid=False, showline=False)
            fig15.update_yaxes(showgrid=False, showline=False)
            st.plotly_chart(fig15, use_container_width=True)

    spacer("xl")

    # ══════════════════════════════════════════════════════════════════════════
    # 10 — OPPORTUNITY OWNER ANALYSIS
    # ══════════════════════════════════════════════════════════════════════════
    st.markdown('<div class="sec">10 · Opportunity Owner Performance</div>', unsafe_allow_html=True)
    spacer("md")

    ow = fdf.groupby("Opportunity Owner").agg(Value=("Opportunity PAR","sum"), Count=("Opportunity Name","count"),
        Avg=("Opportunity PAR","mean"), AvgDur=("Stage Duration","mean")).reset_index().sort_values("Value", ascending=False)

    o8a, _, o8b = st.columns([0.53, 0.06, 0.41])
    with o8a:
        st.markdown('<p class="so">Top opportunity owners by pipeline value</p>', unsafe_allow_html=True)
        ow_top = ow.head(10).sort_values("Value", ascending=True)
        fig16 = go.Figure(go.Bar(
            y=ow_top["Opportunity Owner"], x=ow_top["Value"], orientation="h", marker_color=NY,
            text=[f"  {fc(v)} ({c})" for v,c in zip(ow_top["Value"],ow_top["Count"])],
            textposition="outside", textfont=dict(size=10, color=G800),
        ))
        pl(fig16, h=max(360, 42*len(ow_top)))
        fig16.update_layout(showlegend=False, xaxis=dict(visible=False), yaxis=dict(tickfont=dict(size=10)))
        fig16.update_xaxes(showgrid=False, showline=False)
        fig16.update_yaxes(showgrid=False, showline=False)
        st.plotly_chart(fig16, use_container_width=True)

    with o8b:
        st.markdown('<p class="so">Owner performance table</p>', unsafe_allow_html=True)
        od = ow.rename(columns={"Opportunity Owner":"Owner","Count":"Opps","Value":"Pipeline","Avg":"Avg Deal","AvgDur":"Avg Days"}).copy()
        od["Pipeline"] = od["Pipeline"].apply(lambda x: f"${x:,.0f}")
        od["Avg Deal"] = od["Avg Deal"].apply(lambda x: f"${x:,.0f}")
        od["Avg Days"] = od["Avg Days"].apply(lambda x: f"{x:.0f}")
        st.dataframe(od, use_container_width=True, height=max(320, 38*len(ow_top)), hide_index=True)

    spacer("xl")

    # ══════════════════════════════════════════════════════════════════════════
    # 11 — FULL PIPELINE TABLE
    # ══════════════════════════════════════════════════════════════════════════
    st.markdown('<div class="sec">11 · Full Pipeline Detail</div>', unsafe_allow_html=True)
    spacer("md")

    tbl = fdf[["Stage","Status","Product","Account Name","Opportunity Name","Solution Resource","Opportunity Owner",
               "Main Primary Service","Opportunity PAR","Stage Duration","Close Date Display",
               "Received Display","Closed Display","Notes"]].copy()
    tbl = tbl.rename(columns={"Close Date Display":"Close Date", "Received Display":"Received by Solutions",
                               "Closed Display":"Closed by Solutions"})
    tbl = tbl.sort_values(["Stage","Opportunity PAR"], ascending=[True,False])
    st.dataframe(
        tbl.style.format({"Opportunity PAR":"${:,.0f}"}),
        use_container_width=True, height=min(560, 38*len(tbl)+38), hide_index=True,
        column_config={
            "Account Name": st.column_config.TextColumn("Customer", width="medium"),
            "Opportunity Name": st.column_config.TextColumn("Opportunity", width="large"),
            "Main Primary Service": st.column_config.TextColumn("Service", width="medium"),
            "Status": st.column_config.TextColumn("Status", width="small"),
            "Product": st.column_config.TextColumn("Product", width="small"),
        },
    )

    spacer("xl")

    # ══════════════════════════════════════════════════════════════════════════
    # 12 — EXECUTIVE SUMMARY
    # ══════════════════════════════════════════════════════════════════════════
    st.markdown('<div class="sec">12 · Executive Summary</div>', unsafe_allow_html=True)
    spacer("md")

    top_c  = fdf.groupby("Account Name")["Opportunity PAR"].sum().idxmax() if n_opp else "N/A"
    top_cv = fdf.groupby("Account Name")["Opportunity PAR"].sum().max() if n_opp else 0
    top_sv = fdf.groupby("Main Primary Service")["Opportunity PAR"].sum().idxmax() if n_opp else "N/A"
    n_des  = len(fdf[fdf["Stage"]=="Solutions Design"])
    n_pro  = len(fdf[fdf["Stage"].str.contains("Proposal|Price", case=False, na=False)])
    n_neg  = len(fdf[fdf["Stage"]=="Negotiations"])
    top5p  = cu.head(5)["Value"].sum()/cu["Value"].sum()*100 if len(cu) and cu["Value"].sum()>0 else 0
    reg_lead = f"with <b>{top_reg}</b> leading at <b>{fc(rg.iloc[0]['Value'])}</b>" if len(rg) else ""
    prod_tagged = fdf[fdf["Product"]!="General"]
    n_prod_tagged = len(prod_tagged)
    prod_list = ", ".join(fdf[fdf["Product"]!="General"]["Product"].unique()) if n_prod_tagged else "None tagged"

    st.markdown(f"""<div class="es">
The Solutions team is currently managing <b>{n_opp} active opportunities</b> representing
a total pipeline value of <b>{fc(total)}</b> across <b>{n_cust} unique customers</b>
and <b>{n_svc} service categories</b>.
<br><br>
<b>Stage composition:</b> {n_des} opportunities are in Solutions Design ({pct(n_des,n_opp)}),
{n_pro} in Proposal/Price Quote, {n_neg} in Negotiations, and {len(lost)} were Closed/Lost
({fc(lost['Opportunity PAR'].sum())} lost value).
<br><br>
<b>Status overview:</b> Of all opportunities, <b>{n_working} are actively being worked</b>,
{n_pending} are pending action, and {n_unassigned} remain unassigned.
This signals capacity for re-prioritization across the Solutions team.
<br><br>
<b>Solutions velocity:</b> <b>{n_received} of {n_opp}</b> opportunities ({pct_received}) have been
formally received by Solutions, and <b>{n_closed}</b> have been closed.
Average cycle time (Received → Closed) is <b>{avg_cycle}</b>.
<br><br>
<b>Product tagging:</b> <b>{n_prod_tagged} of {n_opp}</b> opportunities have a product classification
({prod_list}). Expanding product tagging will improve pipeline segmentation and forecasting accuracy.
<br><br>
<b>Customer concentration:</b> The top 5 accounts represent <b>{top5p:.0f}%</b> of pipeline value,
led by <b>{top_c}</b> at {fc(top_cv)}. This is a notable concentration risk.
The dominant service type is <b>{top_sv}</b>.
<br><br>
<b>Regional coverage:</b> Pipeline spans <b>{fdf['Owner Role'].nunique()} regions</b>
{reg_lead}.
<b>{fdf['Solution Resource'].nunique()} Solution Resources</b> are actively engaged.
<br><br>
<b>Attention items:</b> <b>{len(past_due)} opportunities</b> ({fc(past_due['Opportunity PAR'].sum())})
have close dates that have already passed.
<b>{len(aging_60)} opportunities</b> have been in their current stage for over 60 days.
Average stage duration stands at <b>{avg_dur:.0f} days</b>.
</div>""", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════════
#  MASTERFILE MANAGER
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "Masterfile Manager":

    st.markdown(f"""
    <div class="wf">
        <b>Workflow</b><br><br>
        <span class="ws">1</span> Download fresh data from Salesforce<br>
        <span class="ws">2</span> Upload here — the app <b>merges</b> new SF data into the Masterfile<br>
        <span class="ws">3</span> Salesforce fields refresh · <b>Team columns preserved</b> (Solutions Notes, Tasks, Action Items, Comments)<br>
        <span class="ws">4</span> New columns tracked: <b>Status</b>, <b>Received by Solutions</b>, <b>Closed by Solutions</b>, <b>Product</b><br>
        <span class="ws">5</span> Edit team columns inline below<br>
        <span class="ws">6</span> Download the updated Masterfile
    </div>
    """, unsafe_allow_html=True)

    spacer("lg")

    st.markdown('<div class="sec">Upload New Salesforce Export to Merge</div>', unsafe_allow_html=True)
    mf = st.file_uploader("Upload new SF export. Team columns will be preserved.", type=["xlsx","xls","csv"], key="mu")
    if mf:
        raw = pd.read_csv(mf) if mf.name.endswith(".csv") else pd.read_excel(mf)
        new_sf = clean_upload(raw)
        merged, stats = merge_masterfile(st.session_state.master.copy(), new_sf)
        st.session_state.master = merged
        st.success(f"Merge complete — **{stats['updated']}** updated · **{stats['added']}** added · **{stats['removed']}** flagged · **{stats['total']}** total")
        st.rerun()

    spacer("md")

    st.markdown('<div class="sec">Masterfile — Editable</div>', unsafe_allow_html=True)
    st.caption("Salesforce columns are locked. Edit the four team columns (teal-highlighted in Excel download).")

    edf = st.session_state.master.copy()
    for c in TEAM_COLS:
        if c not in edf.columns: edf[c] = ""

    # Format date columns for display
    for dc in ["Close Date", "Received by Solutions", "Closed by Solutions"]:
        if dc in edf.columns:
            edf[dc] = pd.to_datetime(edf[dc], errors="coerce").apply(fmt_date)

    edited = st.data_editor(
        edf, use_container_width=True, height=min(620, 38*len(edf)+38), num_rows="dynamic",
        column_config={
            "Stage": st.column_config.TextColumn("Stage", disabled=True),
            "Account Name": st.column_config.TextColumn("Customer", disabled=True),
            "Opportunity Name": st.column_config.TextColumn("Opportunity", disabled=True, width="large"),
            "Opportunity PAR": st.column_config.NumberColumn("PAR ($)", format="$%d", disabled=True),
            "Stage Duration": st.column_config.NumberColumn("Days", disabled=True),
            "Close Date": st.column_config.TextColumn("Close Date", disabled=True),
            "Owner Role": st.column_config.TextColumn("Region", disabled=True),
            "Opportunity Owner": st.column_config.TextColumn("Opp Owner", disabled=True),
            "Solution Resource": st.column_config.TextColumn("Sol. Resource", disabled=True),
            "Main Primary Service": st.column_config.TextColumn("Service", disabled=True),
            "Notes": st.column_config.TextColumn("SF Notes", disabled=True),
            "Status": st.column_config.TextColumn("Status", disabled=True),
            "Received by Solutions": st.column_config.TextColumn("Received", disabled=True),
            "Closed by Solutions": st.column_config.TextColumn("Closed", disabled=True),
            "Product": st.column_config.TextColumn("Product", disabled=True),
            "Solutions Notes": st.column_config.TextColumn("Solutions Notes", width="large"),
            "Tasks": st.column_config.TextColumn("Tasks", width="large"),
            "Action Items": st.column_config.TextColumn("Action Items", width="large"),
            "Comments / Results": st.column_config.TextColumn("Comments / Results", width="large"),
        },
        hide_index=True, key="editor",
    )

    if st.button("Save edits", type="primary"):
        st.session_state.master = edited.copy()
        st.success("Edits saved to session.")

    spacer("lg")
    st.markdown("---")
    spacer("md")
    d1, d2, _ = st.columns([1,1,2])
    with d1:
        st.download_button("Download Masterfile (.xlsx)", data=to_excel(st.session_state.master),
            file_name=f"Solutions_Masterfile_{datetime.now().strftime('%Y-%m-%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    with d2:
        st.download_button("Download Masterfile (.csv)", data=st.session_state.master.to_csv(index=False).encode(),
            file_name=f"Solutions_Masterfile_{datetime.now().strftime('%Y-%m-%d')}.csv", mime="text/csv")

# Footer
spacer("lg")
st.markdown(f'<div class="ft">MARKEN · UPS HEALTHCARE PRECISION LOGISTICS &nbsp;|&nbsp; SOLUTIONS TEAM PIPELINE REPORT &nbsp;|&nbsp; CONFIDENTIAL</div>', unsafe_allow_html=True)
